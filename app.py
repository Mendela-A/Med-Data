import click
from flask import Flask, render_template, redirect, url_for, request, flash, send_file, jsonify
from flask_migrate import Migrate
from flask_caching import Cache
from config import Config
from models import db, User, Record, bcrypt, log_action, Department, init_db_events
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from sqlalchemy.orm import joinedload
from sqlalchemy import func

login_manager = LoginManager()
login_manager.login_view = 'login'

# Initialize cache
cache = Cache()

def create_app(config_class=Config):
    app = Flask(__name__, template_folder='templates', static_folder='static')
    app.config.from_object(config_class)

    # setup logging: prefer stdout (good for Docker); enable file logging with LOG_TO_FILE=1
    import logging
    import os, sys
    log_to_file = os.environ.get('LOG_TO_FILE') == '1'
    if log_to_file:
        from logging.handlers import RotatingFileHandler
        if not os.path.exists('logs'):
            os.makedirs('logs')
        try:
            file_handler = RotatingFileHandler('logs/app.log', maxBytes=10240, backupCount=3)
            file_handler.setFormatter(logging.Formatter('%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'))
            file_handler.setLevel(logging.INFO)
            app.logger.addHandler(file_handler)
        except Exception:
            # fallback to stderr if file logging cannot be configured
            app.logger.addHandler(logging.StreamHandler(sys.stderr))
            app.logger.warning('Could not configure file logging; logs will be sent to stderr')
    else:
        # In containers, it's best to log to stdout
        app.logger.addHandler(logging.StreamHandler(sys.stdout))
    app.logger.setLevel(logging.INFO)
    app.logger.info('Application startup')

    db.init_app(app)
    bcrypt.init_app(app)
    login_manager.init_app(app)
    Migrate(app, db)
    init_db_events(app)

    # Initialize cache with simple in-memory storage
    cache.init_app(app, config={
        'CACHE_TYPE': 'SimpleCache',
        'CACHE_DEFAULT_TIMEOUT': 900  # 15 minutes default
    })

    # Database maintenance: run ANALYZE periodically to update query planner statistics
    @app.before_request
    def optimize_database():
        """Periodically optimize database statistics for better query planning"""
        import random
        # Run ANALYZE with 1% probability on each request (roughly once per 100 requests)
        if random.random() < 0.01:
            try:
                db.session.execute(db.text("ANALYZE"))
                db.session.commit()
            except Exception:
                pass  # Silent fail - optimization is not critical

    # Ensure data directory exists when using a local sqlite file
    db_uri = app.config.get('SQLALCHEMY_DATABASE_URI', '')
    if db_uri.startswith('sqlite:///'):
        db_path = db_uri.replace('sqlite:///', '')
        db_dir = os.path.dirname(db_path)
        if db_dir and not os.path.exists(db_dir):
            try:
                os.makedirs(db_dir, exist_ok=True)
            except Exception:
                app.logger.warning(f'Could not create directory for sqlite DB: {db_dir}')

    @login_manager.user_loader
    def load_user(user_id):
        return User.query.get(int(user_id))

    from functools import wraps
    def role_required(role):
        """Decorator to require a specific role or admin."""
        def decorator(f):
            @wraps(f)
            def decorated_function(*args, **kwargs):
                if not current_user.is_authenticated:
                    return redirect(url_for('login'))
                # admin has all rights
                if getattr(current_user, 'role', None) != role and getattr(current_user, 'role', None) != 'admin':
                    flash('Доступ заборонено', 'danger')
                    return redirect(url_for('index'))
                return f(*args, **kwargs)
            return decorated_function
        return decorator

    # Cached helper functions for dropdown values
    @cache.memoize(timeout=900)  # Cache for 15 minutes
    def get_distinct_statuses():
        """Get distinct discharge statuses from database"""
        return [s[0] for s in db.session.query(Record.discharge_status).distinct()
                .filter(Record.discharge_status != None)
                .order_by(Record.discharge_status).all()]

    @cache.memoize(timeout=900)  # Cache for 15 minutes
    def get_distinct_physicians():
        """Get distinct treating physicians from database"""
        return [p[0] for p in db.session.query(Record.treating_physician).distinct()
                .filter(Record.treating_physician != None)
                .order_by(Record.treating_physician).all()]

    @cache.memoize(timeout=900)  # Cache for 15 minutes
    def get_distinct_departments():
        """Get departments from Department table"""
        return [d.name for d in Department.query.order_by(Department.name).all()]

    def clear_dropdown_cache():
        """Clear all dropdown caches - call after adding/editing records"""
        cache.delete_memoized(get_distinct_statuses)
        cache.delete_memoized(get_distinct_physicians)
        cache.delete_memoized(get_distinct_departments)

    @app.route('/')
    @login_required
    def index():
        from datetime import datetime, timezone, timedelta
        # Use Kyiv timezone (UTC+2, or UTC+3 during DST) for correct month detection
        # Simple approach: use UTC+2 as base (covers most of the year)
        kyiv_tz = timezone(timedelta(hours=2))
        now = datetime.now(kyiv_tz)

        # support a toggle to show all months
        show_all = request.args.get('all_months', '').lower() in ('1', 'true', 'yes')

        # allow explicit month/year selection via query params or HTML5 month input (YYYY-MM format)
        month_input = request.args.get('month_filter', '').strip()
        selected_month = None
        selected_year = None

        try:
            if month_input:
                # Parse HTML5 month input format: YYYY-MM
                parts = month_input.split('-')
                if len(parts) == 2:
                    selected_year = int(parts[0])
                    selected_month = int(parts[1])
                    if 1 <= selected_month <= 12:
                        start = datetime(selected_year, selected_month, 1)
                        if selected_month == 12:
                            end = datetime(selected_year + 1, 1, 1)
                        else:
                            end = datetime(selected_year, selected_month + 1, 1)
                    else:
                        raise ValueError()
                else:
                    raise ValueError()
            else:
                # Default to current month
                start = datetime(now.year, now.month, 1)
                selected_year = now.year
                selected_month = now.month
                if now.month == 12:
                    end = datetime(now.year + 1, 1, 1)
                else:
                    end = datetime(now.year, now.month + 1, 1)
        except Exception:
            # fallback to current month
            start = datetime(now.year, now.month, 1)
            selected_year = now.year
            selected_month = now.month
            if now.month == 12:
                end = datetime(now.year + 1, 1, 1)
            else:
                end = datetime(now.year, now.month + 1, 1)

        # base query (by default for current month, unless show_all)
        q = Record.query.options(joinedload(Record.creator))
        if not show_all:
            # show records discharged in the current month by date_of_discharge
            q = q.filter(
                Record.date_of_discharge != None,
                Record.date_of_discharge >= start.date(),
                Record.date_of_discharge < end.date(),
            )

        # --- filtering from query params ---
        selected_status = request.args.get('discharge_status', '').strip()
        selected_physician = request.args.get('treating_physician', '').strip()
        selected_department = request.args.get('discharge_department', '').strip()
        history_q = request.args.get('history', '').strip()
        full_name_q = request.args.get('full_name', '').strip()
        has_death_date = request.args.get('has_death_date', '').lower() in ('1', 'true', 'yes')

        conditions = []
        if selected_status:
            conditions.append(Record.discharge_status == selected_status)
        if selected_physician:
            conditions.append(Record.treating_physician == selected_physician)
        if selected_department:
            conditions.append(Record.discharge_department == selected_department)
        if history_q:
            conditions.append(Record.history.contains(history_q))
        if full_name_q:
            conditions.append(Record.full_name.ilike(f'%{full_name_q}%'))
        if has_death_date:
            conditions.append(Record.date_of_death != None)
        if conditions:
            q = q.filter(*conditions)

        # values for dropdowns (cached)
        statuses = get_distinct_statuses()
        physicians = get_distinct_physicians()
        departments = get_distinct_departments()

        # Pagination
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 100, type=int)  # Default 100 records per page
        per_page = min(per_page, 200)  # Max 200 per page to prevent abuse

        # Get paginated results
        pagination = q.order_by(Record.date_of_discharge.desc(), Record.created_at.desc()).paginate(
            page=page,
            per_page=per_page,
            error_out=False
        )
        records = pagination.items
        count = pagination.total

        # Count by discharge status - optimized with database-level aggregation
        # Build query for counting with same filters EXCEPT discharge_status
        count_query = db.session.query(
            Record.discharge_status,
            func.count(Record.id).label('count')
        )

        # Apply base filters (month/show_all)
        if not show_all:
            count_query = count_query.filter(
                Record.date_of_discharge != None,
                Record.date_of_discharge >= start.date(),
                Record.date_of_discharge < end.date(),
            )

        # Apply all filters EXCEPT discharge_status (since we're grouping by it)
        if selected_physician:
            count_query = count_query.filter(Record.treating_physician == selected_physician)
        if selected_department:
            count_query = count_query.filter(Record.discharge_department == selected_department)
        if history_q:
            count_query = count_query.filter(Record.history.contains(history_q))
        if full_name_q:
            count_query = count_query.filter(Record.full_name.ilike(f'%{full_name_q}%'))
        if has_death_date:
            count_query = count_query.filter(Record.date_of_death != None)

        # Group by status and execute
        status_counts = count_query.group_by(Record.discharge_status).all()

        # Extract counts from results (creates a dict: {status: count})
        status_count_dict = {status: cnt for status, cnt in status_counts if status}
        count_discharged = status_count_dict.get('Виписаний', 0)
        count_processing = status_count_dict.get('Опрацьовується', 0)
        count_violations = status_count_dict.get('Порушені вимоги', 0)

        # Format month for HTML5 input (YYYY-MM)
        month_filter_value = f"{selected_year}-{selected_month:02d}" if selected_year and selected_month else ""

        return render_template('dashboard.html', records=records, pagination=pagination, statuses=statuses, physicians=physicians, departments=departments, selected_status=selected_status, selected_physician=selected_physician, selected_department=selected_department, history_q=history_q, full_name_q=full_name_q, count=count, month_filter_value=month_filter_value, selected_year=selected_year, selected_month=selected_month, show_all=show_all, has_death_date=has_death_date, count_discharged=count_discharged, count_processing=count_processing, count_violations=count_violations)

    @app.route('/export-page')
    @role_required('editor')
    def export_page():
        # Get dropdown values for filters (cached)
        physicians = get_distinct_physicians()
        departments = get_distinct_departments()
        return render_template('export.html', physicians=physicians, departments=departments)

    @app.route('/export', methods=['POST'])
    @role_required('editor')
    def export():
        # editors and admins can export records within a date range (by date_of_discharge)
        from datetime import datetime
        from io import BytesIO
        try:
            from_str = request.form.get('from_date', '').strip()
            to_str = request.form.get('to_date', '').strip()
            if not from_str or not to_str:
                flash('Будь ласка, вкажіть обидві дати (з та по) для експорту', 'warning')
                return redirect(url_for('index'))
            from_d = datetime.strptime(from_str, '%Y-%m-%d').date()
            to_d = datetime.strptime(to_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Невірний формат дати для експорту (використовуйте вибір дати)', 'warning')
            return redirect(url_for('index'))

        if from_d > to_d:
            flash('Дата "з" не може бути пізніше дати "по"', 'warning')
            return redirect(url_for('index'))

        # Build query
        q = Record.query.filter(
            Record.date_of_discharge != None,
            Record.date_of_discharge >= from_d,
            Record.date_of_discharge <= to_d,
        )
        # apply active filters if provided (discharge_status, treating_physician, discharge_department, history, full_name)
        discharge_status = request.form.get('discharge_status', '').strip()
        treating_physician = request.form.get('treating_physician', '').strip()
        discharge_department = request.form.get('discharge_department', '').strip()
        history_q = request.form.get('history', '').strip()
        full_name_q = request.form.get('full_name', '').strip()
        if discharge_status:
            q = q.filter(Record.discharge_status == discharge_status)
        if treating_physician:
            q = q.filter(Record.treating_physician == treating_physician)
        if discharge_department:
            q = q.filter(Record.discharge_department == discharge_department)
        if history_q:
            q = q.filter(Record.history.contains(history_q))
        if full_name_q:
            q = q.filter(Record.full_name.ilike(f'%{full_name_q}%'))

        # Count total records first (fast, uses indexes)
        total_count = q.count()

        if total_count == 0:
            flash('Записів не знайдено для обраного діапазону дат та фільтрів', 'warning')
            return redirect(url_for('index'))

        # create excel with openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter
        except Exception:
            flash('Для експорту потрібен пакет openpyxl. Будь ласка, встановіть його.', 'danger')
            return redirect(url_for('index'))

        # Use write_only mode for large exports (>5000 records)
        use_write_only = total_count > 5000
        wb = Workbook(write_only=use_write_only)
        ws = wb.create_sheet() if use_write_only else wb.active
        if not use_write_only:
            ws.title = 'Vipiski'

        # headers: include all columns from the Record model
        headers = ['ID','Дата виписки','ПІБ','Відділення','Лікар','Історія','К днів','Статус виписки','Дата смерті','Коментар','Створив','Створено']

        # Cache user mapping for fast lookups (avoid N+1)
        user_map = {u.id: u.username for u in User.query.all()}

        if use_write_only:
            # Write-only mode: can't style after writing, must style cells inline
            from openpyxl.cell import WriteOnlyCell
            styled_headers = []
            bold = Font(bold=True)
            for h in headers:
                cell = WriteOnlyCell(ws, value=h)
                cell.font = bold
                styled_headers.append(cell)
            ws.append(styled_headers)
        else:
            ws.append(headers)
            # Style headers bold
            bold = Font(bold=True)
            for cell in ws[1]:
                cell.font = bold

        # Batch processing: load records in chunks of 1000
        BATCH_SIZE = 1000
        q = q.order_by(Record.date_of_discharge.desc(), Record.created_at.desc())

        # Calculate column widths from first 100 rows (sampling for performance)
        col_widths = [len(h) for h in headers]
        sample_count = 0

        offset = 0
        while offset < total_count:
            # Fetch batch
            batch = q.offset(offset).limit(BATCH_SIZE).all()

            for r in batch:
                row = [
                    r.id,
                    r.date_of_discharge.strftime('%d.%m.%Y') if r.date_of_discharge else '',
                    r.full_name,
                    r.discharge_department or '',
                    r.treating_physician or '',
                    r.history or '',
                    r.k_days if r.k_days is not None else '',
                    r.discharge_status or '',
                    r.date_of_death.strftime('%d.%m.%Y') if r.date_of_death else '',
                    r.comment or '',
                    user_map.get(r.created_by, r.created_by or ''),
                    r.created_at.strftime('%d.%m.%Y %H:%M') if r.created_at else '',
                ]
                ws.append(row)

                # Sample first 100 rows for column width calculation (not in write_only mode)
                if not use_write_only and sample_count < 100:
                    for i, val in enumerate(row):
                        val_len = len(str(val)) if val is not None else 0
                        if val_len > col_widths[i]:
                            col_widths[i] = val_len
                    sample_count += 1

            offset += BATCH_SIZE

            # Clear session to free memory
            db.session.expire_all()

        # Auto width columns (only in normal mode, not write_only)
        if not use_write_only:
            for i, width in enumerate(col_widths, 1):
                adjusted_width = min(width + 2, 50)
                ws.column_dimensions[get_column_letter(i)].width = adjusted_width

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        # audit export
        try:
            log_action(current_user.id, 'export', 'export', None, f'from={from_d} to={to_d} discharge_status={discharge_status} treating_physician={treating_physician} discharge_department={discharge_department} history={history_q} full_name={full_name_q} count={total_count}')
        except Exception:
            app.logger.exception('Failed to write audit log for export')
        app.logger.info(f'Export by {getattr(current_user, "username", "unknown")}: from {from_d} to {to_d} count={total_count} write_only={use_write_only}')

        filename = f"vipiski_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
        return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/register', methods=['GET', 'POST'])
    def register():
        # Public registration is disabled. Only administrators can create new users via the admin UI.
        flash('Реєстрація вимкнена. Нових користувачів створює лише адміністратор.', 'warning')
        return redirect(url_for('login'))

    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if request.method == 'POST':
            username = request.form.get('username')
            password = request.form.get('password')
            user = User.query.filter_by(username=username).first()
            if user and user.check_password(password):
                login_user(user)
                return redirect(url_for('index'))
            flash('Невірне ім\'я користувача або пароль', 'danger')
            return redirect(url_for('login'))
        return render_template('login.html')

    @app.route('/logout')
    @login_required
    def logout():
        logout_user()
        return redirect(url_for('login'))

    @app.route('/records/add', methods=['GET', 'POST'])
    @role_required('operator')
    def add_record():

        if request.method == 'POST':
            date_str = request.form.get('date_of_discharge', '').strip()
            full_name = request.form.get('full_name', '').strip()
            discharge_department = request.form.get('discharge_department', '').strip()
            treating_physician = request.form.get('treating_physician', '').strip()
            history = request.form.get('history', '').strip()
            k_days = request.form.get('k_days', '').strip()
            date_of_death_str = request.form.get('date_of_death', '').strip()

            # date_of_death (if provided) will indicate death; validate format if present
            # discharge_department is optional (may not exist yet); require other main fields
            if not all([date_str, full_name, treating_physician, history, k_days]):
                flash('Будь ласка, заповніть усі обов\'язкові поля (виключаючи відділення)', 'warning')
                return redirect(url_for('add_record'))

            # validate date format: accept DD.MM.YYYY or YYYY-MM-DD (HTML date input)
            from datetime import datetime
            date_of_discharge = None
            for fmt in ('%d.%m.%Y', '%Y-%m-%d'):
                try:
                    date_of_discharge = datetime.strptime(date_str, fmt).date()
                    break
                except ValueError:
                    continue
            if date_of_discharge is None:
                flash('Дата виписки повинна бути у форматі ДД.ММ.РРРР або РРРР-ММ-ДД', 'warning')
                return redirect(url_for('add_record'))

            # validate k_days integer
            try:
                k_days_int = int(k_days)
            except ValueError:
                flash('"К днів" повинно бути цілим числом', 'warning')
                return redirect(url_for('add_record'))

            date_of_death = None
            if date_of_death_str:
                date_of_death = None
                for fmt in ('%d.%m.%Y', '%Y-%m-%d'):
                    try:
                        date_of_death = datetime.strptime(date_of_death_str, fmt).date()
                        break
                    except ValueError:
                        continue
                if date_of_death is None:
                    flash('Дата смерті повинна бути у форматі ДД.ММ.РРРР або РРРР-ММ-ДД', 'warning')
                    return redirect(url_for('add_record'))

                # date_of_death cannot be earlier than date_of_discharge
                if date_of_death < date_of_discharge:
                    flash('Дата смерті не може бути раніше дати виписки', 'warning')
                    return redirect(url_for('add_record'))

            # Automatically set status to "Опрацьовується" for operator-created records
            discharge_status = 'Опрацьовується'

            r = Record(
                date_of_discharge=date_of_discharge,
                full_name=full_name,
                discharge_department=discharge_department or None,
                treating_physician=treating_physician,
                history=history,
                k_days=k_days_int,
                discharge_status=discharge_status,
                date_of_death=date_of_death,
                created_by=current_user.id,
                updated_by=current_user.id
            )
            db.session.add(r)
            db.session.commit()
            # Clear dropdown cache after adding new record
            clear_dropdown_cache()
            try:
                log_action(current_user.id, 'record.create', 'record', r.id, f'full_name={r.full_name}')
            except Exception:
                app.logger.exception('Failed to write audit log for record.create')
            app.logger.info(f'Record created: {r.id} by {current_user.username}')
            flash(f'Запис #{r.id} ({r.full_name}) успішно додано', 'success')
            # preserve filters if sent with form (check prefixed filter_* fields to avoid shadowing actual inputs)
            params = {}
            for k in ('discharge_status', 'treating_physician', 'history'):
                v = request.form.get(f'filter_{k}', '').strip()
                if v:
                    params[k] = v
            if request.form.get('filter_has_death_date', '').strip():
                params['has_death_date'] = '1'
            return redirect(url_for('index', **params))

        # GET: pass through any filters so add form can include hidden fields and departments
        departments = Department.query.order_by(Department.name).all()
        # Get distinct physicians for autocomplete (cached)
        physicians = get_distinct_physicians()
        return render_template('add_record.html', selected_status=request.args.get('discharge_status', ''), selected_physician=request.args.get('treating_physician', ''), history_q=request.args.get('history', ''), departments=departments, selected_department=request.args.get('discharge_department', ''), physicians=physicians)

    @app.route('/api/records/add', methods=['POST'])
    @role_required('operator')
    def api_add_record():
        """AJAX endpoint for adding records with support for 'save and add another'"""
        from datetime import datetime

        app.logger.info(f'API add_record called by {current_user.username}')
        app.logger.debug(f'Form data: {dict(request.form)}')

        date_str = request.form.get('date_of_discharge', '').strip()
        full_name = request.form.get('full_name', '').strip()
        discharge_department = request.form.get('discharge_department', '').strip()
        treating_physician = request.form.get('treating_physician', '').strip()
        history = request.form.get('history', '').strip()
        k_days = request.form.get('k_days', '').strip()
        date_of_death_str = request.form.get('date_of_death', '').strip()
        comment = request.form.get('comment', '').strip()

        # Validate required fields
        if not all([date_str, full_name, treating_physician, history, k_days]):
            return jsonify({'success': False, 'error': 'Будь ласка, заповніть усі обов\'язкові поля'}), 400

        # Parse date_of_discharge
        date_of_discharge = None
        for fmt in ('%d.%m.%Y', '%Y-%m-%d'):
            try:
                date_of_discharge = datetime.strptime(date_str, fmt).date()
                break
            except ValueError:
                continue
        if date_of_discharge is None:
            return jsonify({'success': False, 'error': 'Невірний формат дати виписки'}), 400

        # Validate k_days
        try:
            k_days_int = int(k_days)
        except ValueError:
            return jsonify({'success': False, 'error': '"К днів" повинно бути цілим числом'}), 400

        # Parse date_of_death if provided
        date_of_death = None
        if date_of_death_str:
            for fmt in ('%d.%m.%Y', '%Y-%m-%d'):
                try:
                    date_of_death = datetime.strptime(date_of_death_str, fmt).date()
                    break
                except ValueError:
                    continue
            if date_of_death is None:
                return jsonify({'success': False, 'error': 'Невірний формат дати смерті'}), 400

            if date_of_death < date_of_discharge:
                return jsonify({'success': False, 'error': 'Дата смерті не може бути раніше дати виписки'}), 400

        # Create record with status "Опрацьовується"
        r = Record(
            date_of_discharge=date_of_discharge,
            full_name=full_name,
            discharge_department=discharge_department or None,
            treating_physician=treating_physician,
            history=history,
            k_days=k_days_int,
            discharge_status='Опрацьовується',
            date_of_death=date_of_death,
            comment=comment or None,
            created_by=current_user.id,
            updated_by=current_user.id
        )

        try:
            db.session.add(r)
            db.session.commit()

            # Clear dropdown cache
            clear_dropdown_cache()

            # Audit log
            try:
                log_action(current_user.id, 'record.create', 'record', r.id, f'full_name={r.full_name}')
            except Exception:
                app.logger.exception('Failed to write audit log for record.create')

            app.logger.info(f'Record created via AJAX: {r.id} by {current_user.username}')

            return jsonify({
                'success': True,
                'record_id': r.id,
                'full_name': r.full_name,
                'message': f'Запис "{r.full_name}" успішно додано'
            })

        except Exception as e:
            db.session.rollback()
            app.logger.exception('Failed to create record via AJAX')
            return jsonify({'success': False, 'error': 'Помилка при збереженні запису'}), 500

    @app.route('/records/<int:record_id>/edit', methods=['GET', 'POST'])
    @role_required('editor')
    def edit_record(record_id):

        r = Record.query.get_or_404(record_id)

        if request.method == 'POST':
            # gather form fields
            date_str = request.form.get('date_of_discharge', '').strip()
            full_name = request.form.get('full_name', '').strip()
            discharge_department = request.form.get('discharge_department', '').strip()
            treating_physician = request.form.get('treating_physician', '').strip()
            history = request.form.get('history', '').strip()
            k_days = request.form.get('k_days', '').strip()
            discharge_status = request.form.get('discharge_status', '').strip()
            date_of_death_str = request.form.get('date_of_death', '').strip()
            comment = request.form.get('comment', '').strip()

            # validate presence of main fields (status is optional; only required when it is 'Помер')
            if not all([date_str, full_name, discharge_department, treating_physician, history, k_days, discharge_status]):
                flash('Будь ласка, заповніть усі обов\'язкові поля', 'warning')
                return redirect(url_for('edit_record', record_id=record_id))

            from datetime import datetime
            date_of_discharge = None
            for fmt in ('%d.%m.%Y', '%Y-%m-%d'):
                try:
                    date_of_discharge = datetime.strptime(date_str, fmt).date()
                    break
                except ValueError:
                    continue
            if date_of_discharge is None:
                flash('Дата виписки повинна бути у форматі ДД.ММ.РРРР або РРРР-ММ-ДД', 'warning')
                return redirect(url_for('edit_record', record_id=record_id))

            try:
                k_days_int = int(k_days)
            except ValueError:
                flash('"К днів" повинно бути цілим числом', 'warning')
                return redirect(url_for('edit_record', record_id=record_id))

            # date_of_death handling
            date_of_death = None

            if date_of_death_str:
                date_of_death = None
                for fmt in ('%d.%m.%Y', '%Y-%m-%d'):
                    try:
                        date_of_death = datetime.strptime(date_of_death_str, fmt).date()
                        break
                    except ValueError:
                        continue
                if date_of_death is None:
                    flash('Дата смерті повинна бути у форматі ДД.ММ.РРРР або РРРР-ММ-ДД', 'warning')
                    return redirect(url_for('edit_record', record_id=record_id))

                if date_of_death < date_of_discharge:
                    flash('Дата смерті не може бути раніше дати виписки', 'warning')
                    return redirect(url_for('edit_record', record_id=record_id))

            # apply changes
            r.date_of_discharge = date_of_discharge
            r.full_name = full_name
            r.discharge_department = discharge_department
            r.treating_physician = treating_physician
            r.history = history
            r.k_days = k_days_int
            r.discharge_status = discharge_status
            r.date_of_death = date_of_death
            r.comment = comment
            r.updated_by = current_user.id
            r.updated_at = datetime.utcnow()

            db.session.commit()
            # Clear dropdown cache after editing record
            clear_dropdown_cache()
            try:
                log_action(current_user.id, 'record.update', 'record', r.id, f'full_name={r.full_name}')
            except Exception:
                app.logger.exception('Failed to write audit log for record.update')
            app.logger.info(f'Record updated: {r.id} by {current_user.username}')
            flash(f'Запис #{r.id} ({r.full_name}) успішно оновлено', 'success')
            params = {}
            for k in ('discharge_status', 'treating_physician', 'history'):
                v = request.form.get(f'filter_{k}', '').strip()
                if v:
                    params[k] = v
            if request.form.get('filter_has_death_date', '').strip():
                params['has_death_date'] = '1'
            # Add anchor to scroll to edited record
            return redirect(url_for('index', **params, _anchor=f'record-{r.id}'))

        # GET -> render form with record data (pass filters through if present) and departments
        departments = Department.query.order_by(Department.name).all()
        # Get distinct physicians for autocomplete (cached)
        physicians = get_distinct_physicians()
        return render_template('edit_record.html', r=r, selected_status=request.args.get('discharge_status', ''), selected_physician=request.args.get('treating_physician', ''), history_q=request.args.get('history', ''), departments=departments, physicians=physicians)

    @app.route('/records/<int:record_id>/delete', methods=['POST'])
    @role_required('admin')
    def delete_record(record_id):
        r = Record.query.get_or_404(record_id)
        db.session.delete(r)
        db.session.commit()
        # Clear dropdown cache after deleting record
        clear_dropdown_cache()
        try:
            log_action(current_user.id, 'record.delete', 'record', r.id, f'full_name={r.full_name}')
        except Exception:
            app.logger.exception('Failed to write audit log for record.delete')
        app.logger.info(f'Record deleted: {r.id} by {current_user.username}')
        flash(f'Запис #{r.id} ({r.full_name}) видалено', 'danger')
        # preserve filters from form (if any)
        params = {}
        for k in ('discharge_status', 'treating_physician', 'discharge_department', 'history', 'full_name'):
            v = request.form.get(k, '').strip()
            if v:
                params[k] = v
        if request.form.get('has_death_date', '').strip():
            params['has_death_date'] = '1'
        return redirect(url_for('index', **params))

    # --- Admin: user management ---
    @app.route('/admin/users')
    @role_required('admin')
    def admin_users():
        users = User.query.order_by(User.username).all()
        return render_template('admin_users.html', users=users)

    @app.route('/admin/users/create', methods=['POST'])
    @role_required('admin')
    def admin_create_user():
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        role = request.form.get('role', '').strip() or 'operator'

        if not username or not password:
            flash('Ім\'я користувача та пароль обов\'язкові', 'warning')
            return redirect(url_for('admin_users'))
        if User.query.filter_by(username=username).first():
            flash('Ім\'я користувача вже зайнято', 'warning')
            return redirect(url_for('admin_users'))

        u = User(username=username, role=role)
        u.set_password(password)
        db.session.add(u)
        db.session.commit()
        # audit & log
        try:
            log_action(current_user.id, 'user.create', 'user', u.id, f'role={role}')
        except Exception:
            app.logger.exception('Failed to write audit log for user.create')
        app.logger.info(f'User created: {username} by {current_user.username}')
        flash(f'Користувача {username} ({role}) успішно створено', 'success')
        return redirect(url_for('admin_users'))

    @app.route('/admin/users/<int:user_id>/edit', methods=['GET', 'POST'])
    @role_required('admin')
    def admin_edit_user(user_id):
        u = User.query.get_or_404(user_id)

        if request.method == 'POST':
            username = request.form.get('username', '').strip()
            password = request.form.get('password', '').strip()
            role = request.form.get('role', '').strip()

            if not username:
                flash('Ім\'я користувача обов\'язкове', 'warning')
                return redirect(url_for('admin_edit_user', user_id=user_id))

            # Check if username is taken by another user
            existing = User.query.filter_by(username=username).first()
            if existing and existing.id != user_id:
                flash('Ім\'я користувача вже зайнято', 'warning')
                return redirect(url_for('admin_edit_user', user_id=user_id))

            # Update username
            old_username = u.username
            u.username = username

            # Update password if provided
            if password:
                u.set_password(password)

            # Update role
            if role in ['operator', 'editor', 'admin']:
                u.role = role

            db.session.commit()

            try:
                details = f'username={old_username}->{username}, role={role}'
                if password:
                    details += ', password_changed=True'
                log_action(current_user.id, 'user.update', 'user', u.id, details)
            except Exception:
                app.logger.exception('Failed to write audit log for user.update')
            app.logger.info(f'User updated: {u.username} by {current_user.username}')
            flash(f'Користувача {u.username} успішно оновлено', 'success')
            return redirect(url_for('admin_users'))

        return render_template('edit_user.html', user=u)

    @app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
    @role_required('admin')
    def admin_delete_user(user_id):
        if current_user.id == user_id:
            flash('Ви не можете видалити самого себе', 'danger')
            return redirect(url_for('admin_users'))
        u = User.query.get_or_404(user_id)
        db.session.delete(u)
        db.session.commit()
        try:
            log_action(current_user.id, 'user.delete', 'user', u.id, f'username={u.username}')
        except Exception:
            app.logger.exception('Failed to write audit log for user.delete')
        app.logger.info(f'User deleted: {u.username} by {current_user.username}')
        flash(f'Користувача {u.username} видалено', 'danger')
        return redirect(url_for('admin_users'))

    # Departments management
    @app.route('/admin/departments')
    @role_required('admin')
    def admin_departments():
        departments = Department.query.order_by(Department.name).all()
        return render_template('admin_departments.html', departments=departments)

    @app.route('/admin/departments/create', methods=['POST'])
    @role_required('admin')
    def admin_create_department():
        name = request.form.get('name', '').strip()
        if not name:
            flash('Назва відділення обов\'язкова', 'warning')
            return redirect(url_for('admin_departments'))
        if Department.query.filter_by(name=name).first():
            flash('Відділення з такою назвою вже існує', 'warning')
            return redirect(url_for('admin_departments'))
        d = Department(name=name)
        db.session.add(d)
        db.session.commit()
        # Clear dropdown cache after creating department
        clear_dropdown_cache()
        try:
            log_action(current_user.id, 'department.create', 'department', d.id, f'name={name}')
        except Exception:
            app.logger.exception('Failed to write audit log for department.create')
        app.logger.info(f'Department created: {name} by {current_user.username}')
        flash(f'Відділення "{name}" успішно створено', 'success')
        return redirect(url_for('admin_departments'))

    @app.route('/admin/departments/<int:dept_id>/delete', methods=['POST'])
    @role_required('admin')
    def admin_delete_department(dept_id):
        d = Department.query.get_or_404(dept_id)
        # prevent deletion if department in use
        in_use = Record.query.filter(Record.discharge_department == d.name).count()
        if in_use:
            flash(f'Неможливо видалити відділення "{d.name}" - використовується в {in_use} записах', 'danger')
            return redirect(url_for('admin_departments'))
        db.session.delete(d)
        db.session.commit()
        # Clear dropdown cache after deleting department
        clear_dropdown_cache()
        try:
            log_action(current_user.id, 'department.delete', 'department', d.id, f'name={d.name}')
        except Exception:
            app.logger.exception('Failed to write audit log for department.delete')
        app.logger.info(f'Department deleted: {d.name} by {current_user.username}')
        flash(f'Відділення "{d.name}" видалено', 'danger')
        return redirect(url_for('admin_departments'))

    @app.cli.command('init-db')
    def init_db():
        """Create database tables."""
        db.create_all()
        click.echo('Initialized the database.')

    @app.cli.command('create-admin')
    @click.argument('username')
    @click.argument('password')
    def create_admin(username, password):
        """Create an admin user: flask create-admin <username> <password>"""
        if User.query.filter_by(username=username).first():
            click.echo('User already exists.')
            return
        u = User(username=username, role='admin')
        u.set_password(password)
        db.session.add(u)
        db.session.commit()
        # audit (CLI-created)
        try:
            log_action(None, 'user.create', 'user', u.id, 'created by CLI')
        except Exception:
            app.logger.exception('Failed to write audit log for create-admin')
        app.logger.info(f'Admin user created by CLI: {username}')
        click.echo(f'Created admin user {username}')

    @app.cli.command('backup-db')
    @click.option('--output', '-o', default=None, help='Output file path (default: data/backup_YYYYMMDD_HHMMSS.db)')
    def backup_db(output):
        """Create a safe backup of the SQLite database (works with WAL mode)."""
        import sqlite3
        import shutil
        from datetime import datetime

        db_uri = app.config.get('SQLALCHEMY_DATABASE_URI', '')
        if not db_uri.startswith('sqlite'):
            click.echo('Backup command only works with SQLite databases')
            return

        source_path = db_uri.replace('sqlite:///', '')
        if not os.path.exists(source_path):
            click.echo(f'Database file not found: {source_path}')
            return

        # Generate default output path
        if not output:
            backup_dir = os.path.dirname(source_path)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output = os.path.join(backup_dir, f'backup_{timestamp}.db')

        # Ensure output directory exists
        output_dir = os.path.dirname(output)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)

        try:
            # Use SQLite backup API for safe hot backup
            source_conn = sqlite3.connect(source_path)
            dest_conn = sqlite3.connect(output)

            source_conn.backup(dest_conn)

            source_conn.close()
            dest_conn.close()

            # Get file size
            size_mb = os.path.getsize(output) / (1024 * 1024)
            click.echo(f'Backup created successfully: {output} ({size_mb:.2f} MB)')
            click.echo(f'Date: {datetime.now().strftime("%d.%m.%Y %H:%M:%S")}')
            app.logger.info(f'Database backup created: {output}')

        except Exception as e:
            click.echo(f'Backup failed: {e}')
            app.logger.exception('Database backup failed')

    @app.cli.command('init-db-with-admin')
    @click.option('--username', default='admin', help='Admin username')
    @click.option('--password', default='admin', help='Admin password')
    def init_db_with_admin(username, password):
        """Create database tables and an admin user if not present."""
        db.create_all()
        if not User.query.filter_by(username=username).first():
            u = User(username=username, role='admin')
            u.set_password(password)
            db.session.add(u)
            db.session.commit()
            try:
                log_action(None, 'user.create', 'user', u.id, 'created by init-db-with-admin')
            except Exception:
                app.logger.exception('Failed to write audit log for init-db-with-admin')
            app.logger.info(f'Admin user created during init: {username}')
            click.echo(f'Created admin user {username}')
        click.echo('Initialized the database (with admin).')

    return app


if __name__ == '__main__':
    app = create_app()
    app.run(debug=True)
