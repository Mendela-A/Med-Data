import sys
import os
import sqlite3
import datetime
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"C:\Users\mendela.a\Downloads\04_data.xlsx"
db_path = os.path.join("data", "app.db")

if not os.path.exists(excel_path):
    print(f"Error: Excel file not found at {excel_path}")
    sys.exit(1)

# Connect to SQLite
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

# Get all departments to build a robust mapping
cursor.execute("SELECT id, name, bed_profile_name, row_no FROM departments")
depts = cursor.fetchall()

# Mapping dicts
dept_by_row_no = {}
gyn_dept_id = None

print("Available departments in DB:")
for d in depts:
    d_id, name, profile, row_no = d
    print(f"  ID: {d_id}, Name: '{name}', Profile: '{profile}', RowNo: {row_no}")
    if name == 'Гінекологічне' or profile == 'Пологовий будинок':
        gyn_dept_id = d_id
    if row_no is not None:
        dept_by_row_no[int(row_no)] = d_id

print(f"\nHardcoded Gyn Department ID: {gyn_dept_id}")
print(f"Mapped row numbers: {list(dept_by_row_no.keys())}")

# Load Workbook
print(f"\nLoading Excel workbook: {excel_path}...")
wb = openpyxl.load_workbook(excel_path, data_only=True)
print("Workbook loaded successfully.")

# We will clear the existing April 2026 data in daily_reports
print("\nDeleting existing daily reports for April 2026...")
cursor.execute("DELETE FROM daily_reports WHERE report_date >= '2026-04-01' AND report_date <= '2026-04-30'")
print(f"Deleted {cursor.rowcount} records.")

def to_int(val):
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if not val or val == '':
            return None
    try:
        # Handle cases where excel might return float for integer
        return int(float(val))
    except (ValueError, TypeError):
        return None

now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
inserted_count = 0

# Loop through sheets '01' to '30'
for day_num in range(1, 31):
    sheet_name = f"{day_num:02d}"
    if sheet_name not in wb.sheetnames:
        print(f"Warning: Sheet {sheet_name} not found in Excel, skipping.")
        continue
    
    sheet = wb[sheet_name]
    report_date_str = f"2026-04-{day_num:02d}"
    print(f"Importing sheet '{sheet_name}' for date {report_date_str}...")
    
    # Rows 12 to 40 contain the department data
    for r in range(12, 40):
        name_val = sheet.cell(row=r, column=1).value
        if name_val is None:
            continue
        
        name_val = name_val.strip()
        # Skip totals rows
        if any(term in name_val for term in ['КНП', 'ВСЬОГО', 'Всього', 'Разом', 'Разом по']):
            continue
        
        row_no_val = to_int(sheet.cell(row=r, column=2).value)
        
        # Match department_id
        department_id = None
        if name_val == 'Пологовий будинок':
            department_id = gyn_dept_id
        elif row_no_val in dept_by_row_no:
            department_id = dept_by_row_no[row_no_val]
        else:
            # Fallback by name contains if row number not found or not matched
            for d in depts:
                d_id, name, profile, row_no = d
                if name_val.lower() in name.lower() or (profile and name_val.lower() in profile.lower()):
                    department_id = d_id
                    break
        
        if department_id is None:
            print(f"  Warning: Could not match department for row {r}: '{name_val}' (row_no: {row_no_val})")
            continue
        
        # Extract values
        beds_total = to_int(sheet.cell(row=r, column=3).value)
        beds_renovation = to_int(sheet.cell(row=r, column=4).value)
        patients_start = to_int(sheet.cell(row=r, column=5).value)
        admitted_total = to_int(sheet.cell(row=r, column=6).value)
        admitted_rural = to_int(sheet.cell(row=r, column=7).value)
        admitted_children = to_int(sheet.cell(row=r, column=8).value)
        transferred_in = to_int(sheet.cell(row=r, column=9).value)
        transferred_out = to_int(sheet.cell(row=r, column=10).value)
        discharged_total = to_int(sheet.cell(row=r, column=11).value)
        discharged_to_other = to_int(sheet.cell(row=r, column=12).value)
        deaths = to_int(sheet.cell(row=r, column=13).value)
        patients_end = to_int(sheet.cell(row=r, column=14).value)
        patients_end_rural = to_int(sheet.cell(row=r, column=15).value)
        mothers_with_children = to_int(sheet.cell(row=r, column=16).value)
        free_male = to_int(sheet.cell(row=r, column=17).value)
        free_female = to_int(sheet.cell(row=r, column=18).value)
        
        # Insert statement
        cursor.execute("""
            INSERT INTO daily_reports (
                report_date, department_id, beds_total, beds_renovation, patients_start,
                admitted_total, admitted_rural, admitted_children, transferred_in, transferred_out,
                discharged_total, discharged_to_other, deaths, patients_end, patients_end_rural,
                mothers_with_children, free_male, free_female, created_by, created_at, updated_at,
                admitted_children_rural
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?, NULL)
        """, (
            report_date_str, department_id, beds_total, beds_renovation, patients_start,
            admitted_total, admitted_rural, admitted_children, transferred_in, transferred_out,
            discharged_total, discharged_to_other, deaths, patients_end, patients_end_rural,
            mothers_with_children, free_male, free_female, now_str, now_str
        ))
        inserted_count += 1

# Commit changes
conn.commit()
print(f"\nImport finished. Successfully inserted {inserted_count} daily reports into SQLite database.")

# Quick verification: count again
cursor.execute("SELECT count(*) FROM daily_reports WHERE report_date >= '2026-04-01' AND report_date <= '2026-04-30'")
count_after = cursor.fetchone()[0]
print(f"Total reports in April 2026 after import: {count_after}")

conn.close()
