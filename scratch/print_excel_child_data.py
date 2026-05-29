import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"C:\Users\mendela.a\Downloads\04_data.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

print("Excel values for children's departments in all sheets:")

for day_num in range(1, 31):
    sheet_name = f"{day_num:02d}"
    if sheet_name not in wb.sheetnames:
        continue
    sheet = wb[sheet_name]
    
    for r in range(12, 33):
        name_val = sheet.cell(row=r, column=1).value
        if name_val is None:
            continue
        name_val = name_val.strip()
        
        if any(kd in name_val.lower() for kd in ['дит', 'діти']):
            row_no = sheet.cell(row=r, column=2).value
            beds = sheet.cell(row=r, column=3).value
            start = sheet.cell(row=r, column=5).value
            admitted = sheet.cell(row=r, column=6).value
            discharged = sheet.cell(row=r, column=11).value
            end = sheet.cell(row=r, column=14).value
            
            # Print only if there's any non-None value in movement columns
            print(f"Sheet {sheet_name} | Name: '{name_val}' | RowNo: {row_no} | Beds: {beds} | Start: {start} | Admit: {admitted} | Disch: {discharged} | End: {end}")
