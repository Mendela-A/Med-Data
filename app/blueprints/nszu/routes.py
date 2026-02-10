# app/blueprints/nszu/routes.py
"""
NSZU Corrections routes
"""

from flask import render_template, redirect, url_for, flash, request, current_app, send_file, jsonify
from flask_login import login_required, current_user
from datetime import datetime, timezone
from calendar import monthrange
from io import BytesIO
import locale

from app.extensions import db
from models import NSZUCorrection, User, log_action
from decorators import role_required
from utils import parse_date, parse_numeric
from . import nszu_bp


@nszu_bp.route('')
@role_required('editor', 'viewer')
def nszu_list():
    """List all NSZU corrections with filters"""
    q = NSZUCorrection.query

    # Month filter - default to current month
    month_year_str = request.args.get('month_year', '').strip()
    if month_year_str:
        try:
            year, month = map(int, month_year_str.split('-'))
        except (ValueError, AttributeError):
            year = datetime.now().year
            month = datetime.now().month
    else:
        year = datetime.now().year
        month = datetime.now().month

    # Filter by month
    start_date = datetime(year, month, 1).date()
    last_day = monthrange(year, month)[1]
    end_date = datetime(year, month, last_day).date()

    # Filtering
    selected_status = request.args.get('status', '').strip()
    selected_doctor = request.args.get('doctor', '').strip()
    nszu_record_id_q = request.args.get('nszu_record_id', '').strip()

    conditions = [
        NSZUCorrection.date >= start_date,
        NSZUCorrection.date <= end_date
    ]
    if selected_status:
        conditions.append(NSZUCorrection.status == selected_status)
    if selected_doctor:
        conditions.append(NSZUCorrection.doctor == selected_doctor)
    if nszu_record_id_q:
        conditions.append(NSZUCorrection.nszu_record_id.contains(nszu_record_id_q))

    q = q.filter(*conditions)

    # Get distinct values for filters
    statuses = ['В обробці', 'Опрацьовано', 'Оплачено', 'Не підлягає оплаті']
    doctors = [d[0] for d in db.session.query(NSZUCorrection.doctor).distinct().filter(NSZUCorrection.doctor != None).order_by(NSZUCorrection.doctor).all()]

    # Sorting
    sort_by = request.args.get('sort_by', 'date')
    sort_order = request.args.get('sort_order', 'desc')

    # Map sort columns to model fields
    sort_columns = {
        'id': NSZUCorrection.id,
        'date': NSZUCorrection.date,
        'nszu_record_id': NSZUCorrection.nszu_record_id,
        'doctor': NSZUCorrection.doctor,
        'status': NSZUCorrection.status,
        'fakt_summ': NSZUCorrection.fakt_summ
    }

    # Apply sorting
    if sort_by in sort_columns:
        sort_column = sort_columns[sort_by]
        if sort_order == 'asc':
            q = q.order_by(sort_column.asc())
        else:
            q = q.order_by(sort_column.desc())
        # Secondary sort by created_at for consistency
        q = q.order_by(NSZUCorrection.created_at.desc())
    else:
        # Default sorting
        q = q.order_by(NSZUCorrection.date.desc(), NSZUCorrection.created_at.desc())

    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 100, type=int)
    per_page = min(per_page, 200)

    pagination = q.paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )

    corrections = pagination.items
    count = pagination.total

    # Calculate quick statistics for filtered records
    from sqlalchemy import func
    filtered_stats = db.session.query(
        NSZUCorrection.status,
        func.count(NSZUCorrection.id).label('count'),
        func.sum(NSZUCorrection.fakt_summ).label('total_sum')
    ).filter(*conditions).group_by(NSZUCorrection.status).all()

    status_stats = {stat.status: {'count': stat.count, 'sum': float(stat.total_sum or 0)} for stat in filtered_stats}
    total_filtered_sum = sum(stat['sum'] for stat in status_stats.values())

    # Format current month for display
    try:
        locale.setlocale(locale.LC_TIME, 'uk_UA.UTF-8')
    except locale.Error:
        try:
            locale.setlocale(locale.LC_TIME, 'Ukrainian_Ukraine.1251')
        except locale.Error:
            pass
    current_month = datetime(year, month, 1).strftime('%B %Y')

    # Calculate navigation months
    now = datetime.now()
    current_month_str = f'{now.year:04d}-{now.month:02d}'

    # Previous month
    if month == 1:
        prev_month_str = f'{year-1:04d}-12'
    else:
        prev_month_str = f'{year:04d}-{month-1:02d}'

    # Next month
    if month == 12:
        next_month_str = f'{year+1:04d}-01'
    else:
        next_month_str = f'{year:04d}-{month+1:02d}'

    return render_template('nszu_list.html',
                         corrections=corrections,
                         pagination=pagination,
                         statuses=statuses,
                         doctors=doctors,
                         selected_status=selected_status,
                         selected_doctor=selected_doctor,
                         nszu_record_id_q=nszu_record_id_q,
                         count=count,
                         status_stats=status_stats,
                         total_filtered_sum=total_filtered_sum,
                         selected_year=year,
                         selected_month=month,
                         current_month=current_month,
                         current_month_str=current_month_str,
                         prev_month_str=prev_month_str,
                         next_month_str=next_month_str,
                         sort_by=sort_by,
                         sort_order=sort_order)


@nszu_bp.route('/add', methods=['GET', 'POST'])
@role_required('editor')
def nszu_add():
    """Add new NSZU correction"""
    if request.method == 'POST':
        date_str = request.form.get('date', '').strip()
        nszu_record_id = request.form.get('nszu_record_id', '').strip()
        doctor = request.form.get('doctor', '').strip()
        status = request.form.get('status', 'В обробці').strip()
        detail = request.form.get('detail', '').strip()
        fakt_summ_str = request.form.get('fakt_summ', '').strip()
        comment = request.form.get('comment', '').strip()

        # Validate required fields
        if not all([date_str, nszu_record_id, doctor]):
            flash('Будь ласка, заповніть усі обов\'язкові поля (дата, НСЗУ ID, лікар)', 'warning')
            return redirect(url_for('nszu.nszu_add'))

        # Parse date
        date_obj = parse_date(date_str)
        if date_obj is None:
            flash('Дата повинна бути у форматі ДД.ММ.РРРР або РРРР-ММ-ДД', 'warning')
            return redirect(url_for('nszu.nszu_add'))

        # Parse fakt_summ
        if fakt_summ_str and fakt_summ_str != '-':
            fakt_summ = parse_numeric(fakt_summ_str, default=0.0)
            if fakt_summ is None:
                flash('Фактична сума повинна бути числом', 'warning')
                return redirect(url_for('nszu.nszu_add'))
        else:
            fakt_summ = 0.00

        # Create correction
        correction = NSZUCorrection(
            date=date_obj,
            nszu_record_id=nszu_record_id,
            doctor=doctor,
            status=status,
            detail=detail or None,
            fakt_summ=fakt_summ,
            comment=comment or None,
            created_by=current_user.id,
            updated_by=current_user.id
        )

        db.session.add(correction)
        db.session.commit()

        try:
            log_action(current_user.id, 'nszu.create', 'nszu_correction', correction.id, f'nszu_record_id={nszu_record_id}')
        except Exception:
            current_app.logger.exception('Failed to write audit log for nszu.create')

        current_app.logger.info(f'NSZU correction created: {correction.id} by {current_user.username}')
        flash(f'Запис перевірки НСЗУ #{correction.id} успішно додано', 'success')
        return redirect(url_for('nszu.nszu_list'))

    # GET - render form
    # Get distinct doctors for autocomplete
    doctors = [d[0] for d in db.session.query(NSZUCorrection.doctor).distinct().filter(NSZUCorrection.doctor != None).order_by(NSZUCorrection.doctor).all()]
    statuses = ['В обробці', 'Опрацьовано', 'Оплачено', 'Не підлягає оплаті']

    return render_template('nszu_add.html', doctors=doctors, statuses=statuses)


@nszu_bp.route('/api/add', methods=['POST'])
@role_required('editor')
def api_nszu_add():
    """AJAX endpoint for adding NSZU corrections with support for 'save and add another'"""
    current_app.logger.info(f'API nszu_add called by {current_user.username}')

    date_str = request.form.get('date', '').strip()
    nszu_record_id = request.form.get('nszu_record_id', '').strip()
    doctor = request.form.get('doctor', '').strip()
    status = request.form.get('status', 'В обробці').strip()
    detail = request.form.get('detail', '').strip()
    fakt_summ_str = request.form.get('fakt_summ', '').strip()
    comment = request.form.get('comment', '').strip()

    # Validate required fields
    if not all([date_str, nszu_record_id, doctor]):
        return jsonify({'success': False, 'error': 'Будь ласка, заповніть усі обов\'язкові поля (дата, НСЗУ ID, лікар)'}), 400

    # Parse date
    date_obj = parse_date(date_str)
    if date_obj is None:
        return jsonify({'success': False, 'error': 'Дата повинна бути у форматі ДД.ММ.РРРР або РРРР-ММ-ДД'}), 400

    # Parse fakt_summ
    if fakt_summ_str and fakt_summ_str != '-':
        fakt_summ = parse_numeric(fakt_summ_str, default=0.0)
        if fakt_summ is None:
            return jsonify({'success': False, 'error': 'Фактична сума повинна бути числом'}), 400
    else:
        fakt_summ = 0.00

    # Create correction
    correction = NSZUCorrection(
        date=date_obj,
        nszu_record_id=nszu_record_id,
        doctor=doctor,
        status=status,
        detail=detail or None,
        fakt_summ=fakt_summ,
        comment=comment or None,
        created_by=current_user.id,
        updated_by=current_user.id
    )

    try:
        db.session.add(correction)
        db.session.commit()

        try:
            log_action(current_user.id, 'nszu.create', 'nszu_correction', correction.id, f'nszu_record_id={nszu_record_id}')
        except Exception:
            current_app.logger.exception('Failed to write audit log for nszu.create')

        current_app.logger.info(f'NSZU correction created: {correction.id} by {current_user.username}')
        return jsonify({
            'success': True,
            'message': f'Запис перевірки НСЗУ #{correction.id} успішно додано',
            'correction_id': correction.id
        }), 201

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error creating NSZU correction')
        return jsonify({'success': False, 'error': 'Помилка при збереженні запису'}), 500


@nszu_bp.route('/<int:correction_id>/edit', methods=['GET', 'POST'])
@role_required('editor')
def nszu_edit(correction_id):
    """Edit NSZU correction"""
    correction = NSZUCorrection.query.get_or_404(correction_id)

    if request.method == 'POST':
        # Collect filter params from hidden fields for redirect
        filter_params = {
            'month_year': request.form.get('filter_month_year', ''),
            'status': request.form.get('filter_status', ''),
            'doctor': request.form.get('filter_doctor', ''),
            'nszu_record_id': request.form.get('filter_nszu_record_id', ''),
            'per_page': request.form.get('filter_per_page', ''),
            'sort_by': request.form.get('filter_sort_by', ''),
            'sort_order': request.form.get('filter_sort_order', ''),
        }

        date_str = request.form.get('date', '').strip()
        nszu_record_id = request.form.get('nszu_record_id', '').strip()
        doctor = request.form.get('doctor', '').strip()
        status = request.form.get('status', '').strip()
        detail = request.form.get('detail', '').strip()
        fakt_summ_str = request.form.get('fakt_summ', '').strip()
        comment = request.form.get('comment', '').strip()

        # Validate required fields
        if not all([date_str, nszu_record_id, doctor, status]):
            flash('Будь ласка, заповніть усі обов\'язкові поля', 'warning')
            return redirect(url_for('nszu.nszu_edit', correction_id=correction_id, **filter_params))

        # Parse date
        date_obj = parse_date(date_str)
        if date_obj is None:
            flash('Дата повинна бути у форматі ДД.ММ.РРРР або РРРР-ММ-ДД', 'warning')
            return redirect(url_for('nszu.nszu_edit', correction_id=correction_id, **filter_params))

        # Parse fakt_summ
        if fakt_summ_str and fakt_summ_str != '-':
            fakt_summ = parse_numeric(fakt_summ_str, default=0.0)
            if fakt_summ is None:
                flash('Фактична сума повинна бути числом', 'warning')
                return redirect(url_for('nszu.nszu_edit', correction_id=correction_id, **filter_params))
        else:
            fakt_summ = 0.00

        # Update correction
        correction.date = date_obj
        correction.nszu_record_id = nszu_record_id
        correction.doctor = doctor
        correction.status = status
        correction.detail = detail or None
        correction.fakt_summ = fakt_summ
        correction.comment = comment or None
        correction.updated_by = current_user.id
        correction.updated_at = datetime.now(timezone.utc)

        db.session.commit()

        try:
            log_action(current_user.id, 'nszu.update', 'nszu_correction', correction.id, f'nszu_record_id={nszu_record_id}')
        except Exception:
            current_app.logger.exception('Failed to write audit log for nszu.update')

        current_app.logger.info(f'NSZU correction updated: {correction.id} by {current_user.username}')
        flash(f'Запис перевірки НСЗУ #{correction.id} успішно оновлено', 'success')
        return redirect(url_for('nszu.nszu_list', **filter_params))

    # GET - render form
    # Collect filter params from query string to pass to template
    filter_params = {
        'filter_month_year': request.args.get('month_year', ''),
        'filter_status': request.args.get('status', ''),
        'filter_doctor': request.args.get('doctor', ''),
        'filter_nszu_record_id': request.args.get('nszu_record_id', ''),
        'filter_per_page': request.args.get('per_page', ''),
        'filter_sort_by': request.args.get('sort_by', ''),
        'filter_sort_order': request.args.get('sort_order', ''),
    }

    doctors = [d[0] for d in db.session.query(NSZUCorrection.doctor).distinct().filter(NSZUCorrection.doctor != None).order_by(NSZUCorrection.doctor).all()]
    statuses = ['В обробці', 'Опрацьовано', 'Оплачено', 'Не підлягає оплаті']

    return render_template('nszu_edit.html', correction=correction, doctors=doctors, statuses=statuses, **filter_params)


@nszu_bp.route('/<int:correction_id>/delete', methods=['POST'])
@role_required('admin')
def nszu_delete(correction_id):
    """Delete NSZU correction (admin only)"""
    correction = NSZUCorrection.query.get_or_404(correction_id)
    nszu_id = correction.nszu_record_id

    db.session.delete(correction)
    db.session.commit()

    try:
        log_action(current_user.id, 'nszu.delete', 'nszu_correction', correction_id, f'nszu_record_id={nszu_id}')
    except Exception:
        current_app.logger.exception('Failed to write audit log for nszu.delete')

    current_app.logger.info(f'NSZU correction deleted: {correction_id} by {current_user.username}')
    flash(f'Запис перевірки НСЗУ #{correction_id} видалено', 'danger')
    return redirect(url_for('nszu.nszu_list',
                            month_year=request.form.get('month_year', ''),
                            status=request.form.get('status', ''),
                            doctor=request.form.get('doctor', ''),
                            nszu_record_id=request.form.get('nszu_record_id', ''),
                            per_page=request.form.get('per_page', ''),
                            sort_by=request.form.get('sort_by', ''),
                            sort_order=request.form.get('sort_order', '')))


@nszu_bp.route('/export', methods=['POST'])
@role_required('editor', 'viewer')
def nszu_export():
    """Export NSZU corrections to Excel with filters support"""
    try:
        from_str = request.form.get('from_date', '').strip()
        to_str = request.form.get('to_date', '').strip()
        if not from_str or not to_str:
            flash('Будь ласка, вкажіть обидві дати (з та по) для експорту', 'warning')
            return redirect(url_for('nszu.nszu_list'))
        from_d = datetime.strptime(from_str, '%Y-%m-%d').date()
        to_d = datetime.strptime(to_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Невірний формат дати для експорту', 'warning')
        return redirect(url_for('nszu.nszu_list'))

    if from_d > to_d:
        flash('Дата "з" не може бути пізніше дати "по"', 'warning')
        return redirect(url_for('nszu.nszu_list'))

    # Get optional filters
    status_filter = request.form.get('status', '').strip()
    doctor_filter = request.form.get('doctor', '').strip()
    nszu_id_filter = request.form.get('nszu_record_id', '').strip()

    # Build query with filters
    conditions = [
        NSZUCorrection.date >= from_d,
        NSZUCorrection.date <= to_d,
    ]
    if status_filter:
        conditions.append(NSZUCorrection.status == status_filter)
    if doctor_filter:
        conditions.append(NSZUCorrection.doctor == doctor_filter)
    if nszu_id_filter:
        conditions.append(NSZUCorrection.nszu_record_id.contains(nszu_id_filter))

    q = NSZUCorrection.query.filter(*conditions)

    total_count = q.count()

    if total_count == 0:
        flash('Записів не знайдено для обраного діапазону дат', 'warning')
        return redirect(url_for('nszu.nszu_list'))

    # Create Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, numbers
        from openpyxl.utils import get_column_letter
    except Exception:
        flash('Для експорту потрібен пакет openpyxl', 'danger')
        return redirect(url_for('nszu.nszu_list'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'NSZU'

    # Headers
    headers = ['ID', 'Дата', 'НСЗУ ID', 'Лікар', 'Статус', 'Деталі', 'Факт. сума', 'Коментар', 'Створив', 'Створено', 'Оновив', 'Оновлено']
    ws.append(headers)

    # Style headers
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Get user mapping
    user_map = {u.id: u.username for u in User.query.all()}

    # Add data
    for c in q.order_by(NSZUCorrection.date.desc()).all():
        row = [
            c.id,
            c.date.strftime('%d.%m.%Y') if c.date else '',
            c.nszu_record_id or '',
            c.doctor or '',
            c.status or '',
            c.detail or '',
            float(c.fakt_summ) if c.fakt_summ else 0.00,
            c.comment or '',
            user_map.get(c.created_by, c.created_by or ''),
            c.created_at.strftime('%d.%m.%Y %H:%M') if c.created_at else '',
            user_map.get(c.updated_by, c.updated_by or '') if c.updated_by else '',
            c.updated_at.strftime('%d.%m.%Y %H:%M') if c.updated_at else '',
        ]
        ws.append(row)

    # Format sum column as number with 2 decimal places
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=7).number_format = numbers.FORMAT_NUMBER_00

    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    # Build filename with filters info
    filename_parts = ['nszu', f'{from_d.year}-{from_d.month:02d}']
    if status_filter:
        filename_parts.append(status_filter.replace(' ', '-'))
    filename_parts.append(datetime.now().strftime('%d-%m-%Y'))
    filename = f"{'_'.join(filename_parts)}.xlsx"

    try:
        log_details = f'from={from_d} to={to_d} status={status_filter} doctor={doctor_filter} count={total_count}'
        log_action(current_user.id, 'nszu.export', 'export', None, log_details)
    except Exception:
        current_app.logger.exception('Failed to write audit log for nszu.export')

    current_app.logger.info(f'NSZU export by {current_user.username}: {log_details}')

    return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@nszu_bp.route('/print', methods=['POST'])
@role_required('editor', 'viewer')
def print_nszu():
    """Generate PDF print for NSZU corrections with filters"""
    try:
        from_str = request.form.get('from_date', '').strip()
        to_str = request.form.get('to_date', '').strip()
        if not from_str or not to_str:
            flash('Будь ласка, вкажіть обидві дати для друку', 'warning')
            return redirect(url_for('nszu.nszu_list'))
        from_d = datetime.strptime(from_str, '%Y-%m-%d').date()
        to_d = datetime.strptime(to_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Невірний формат дати для друку', 'warning')
        return redirect(url_for('nszu.nszu_list'))

    if from_d > to_d:
        flash('Дата "з" не може бути пізніше дати "по"', 'warning')
        return redirect(url_for('nszu.nszu_list'))

    # Get optional filters
    status_filter = request.form.get('status', '').strip()
    doctor_filter = request.form.get('doctor', '').strip()
    nszu_id_filter = request.form.get('nszu_record_id', '').strip()

    # Build query with filters
    conditions = [
        NSZUCorrection.date >= from_d,
        NSZUCorrection.date <= to_d,
    ]
    if status_filter:
        conditions.append(NSZUCorrection.status == status_filter)
    if doctor_filter:
        conditions.append(NSZUCorrection.doctor == doctor_filter)
    if nszu_id_filter:
        conditions.append(NSZUCorrection.nszu_record_id.contains(nszu_id_filter))

    q = NSZUCorrection.query.filter(*conditions)
    corrections = q.order_by(NSZUCorrection.date.desc()).all()

    if not corrections:
        flash('Записів не знайдено для обраного діапазону дат', 'warning')
        return redirect(url_for('nszu.nszu_list'))

    # Render HTML template
    html_string = render_template('print_nszu.html',
                                 corrections=corrections,
                                 from_date=from_d,
                                 to_date=to_d,
                                 status_filter=status_filter,
                                 doctor_filter=doctor_filter,
                                 nszu_id_filter=nszu_id_filter,
                                 generated_by=current_user.username,
                                 generated_at=datetime.now())

    # Generate PDF with WeasyPrint
    try:
        from weasyprint import HTML
        pdf = HTML(string=html_string).write_pdf()
    except Exception as e:
        current_app.logger.error(f'PDF generation error: {e}')
        flash('Помилка при генерації PDF', 'danger')
        return redirect(url_for('nszu.nszu_list'))

    # Return PDF file
    bio = BytesIO(pdf)
    bio.seek(0)
    filename = f"nszu_print_{datetime.now().strftime('%d-%m-%Y')}.pdf"

    return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/pdf')
