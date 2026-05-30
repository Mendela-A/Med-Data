# app/blueprints/ambulatory/routes.py
"""
Ambulatory routes - handles CRUD operations, filters, Excel exports, and PDF generation for Ambulatory records.
"""

from flask import render_template, redirect, url_for, flash, request, current_app, send_file, jsonify
from flask_login import login_required, current_user
from datetime import datetime, timezone, timedelta
from io import BytesIO
from sqlalchemy.orm import joinedload
from sqlalchemy import func

from app.extensions import db
from models import AmbulatoryRecord, User, log_action
from decorators import role_required
from utils import (parse_date, clear_dropdown_cache, get_user_map, escape_like,
                   validate_ambulatory_form, get_distinct_ambulatory_statuses,
                   get_distinct_ambulatory_doctors)
from constants import STATUS_PROCESSING, STATUS_DISCHARGED, STATUS_VIOLATIONS
from . import ambulatory_bp


@ambulatory_bp.route('/')
@login_required
def index():
    # Use Kyiv timezone (UTC+2, or UTC+3 during DST)
    kyiv_tz = timezone(timedelta(hours=2))
    now = datetime.now(kyiv_tz)

    # Toggle to show all months
    show_all = request.args.get('all_months', '').lower() in ('1', 'true', 'yes')

    # Parse HTML5 month filter
    month_input = request.args.get('month_filter', '').strip()
    from_date_input = request.args.get('from_date', '').strip()
    to_date_input = request.args.get('to_date', '').strip()
    selected_month = None
    selected_year = None

    try:
        if from_date_input and to_date_input:
            # Date range mode
            from datetime import date as date_type
            fd = date_type.fromisoformat(from_date_input)
            td = date_type.fromisoformat(to_date_input)
            if fd > td:
                fd, td = td, fd
            start = datetime(fd.year, fd.month, fd.day)
            end = datetime(td.year, td.month, td.day) + timedelta(days=1)
            selected_year = fd.year
            selected_month = fd.month
        elif month_input:
            parts = month_input.split('-')
            if len(parts) == 2:
                selected_year = int(parts[0])
                selected_month = int(parts[1])
                if 1 <= selected_month <= 12:
                    start = datetime(selected_year, selected_month, 1)
                    if selected_month == 12:
                        end = datetime(selected_year + 1, 1, 1)
                    else:
                        end = datetime(selected_year, selected_month + 1, 1)
                else:
                    raise ValueError()
            else:
                raise ValueError()
        else:
            # Default to current month
            start = datetime(now.year, now.month, 1)
            selected_year = now.year
            selected_month = now.month
            if now.month == 12:
                end = datetime(now.year + 1, 1, 1)
            else:
                end = datetime(now.year, now.month + 1, 1)
    except Exception:
        start = datetime(now.year, now.month, 1)
        selected_year = now.year
        selected_month = now.month
        if now.month == 12:
            end = datetime(now.year + 1, 1, 1)
        else:
            end = datetime(now.year, now.month + 1, 1)

    # Base query
    q = AmbulatoryRecord.query.options(joinedload(AmbulatoryRecord.creator), joinedload(AmbulatoryRecord.updater))
    if not show_all:
        q = q.filter(
            AmbulatoryRecord.date >= start.date(),
            AmbulatoryRecord.date < end.date(),
        )

    # Apply filters
    selected_status = request.args.get('discharge_status', '').strip()
    selected_doctor = request.args.get('doctor', '').strip()
    full_name_q = request.args.get('full_name', '').strip()
    journal_number_q = request.args.get('journal_number', '').strip()
    diagnosis_q = request.args.get('diagnosis', '').strip()

    conditions = []
    if selected_status:
        conditions.append(AmbulatoryRecord.discharge_status == selected_status)
    if selected_doctor:
        conditions.append(AmbulatoryRecord.doctor == selected_doctor)
    if full_name_q:
        conditions.append(AmbulatoryRecord.full_name.ilike(f'%{escape_like(full_name_q)}%', escape='\\'))
    if journal_number_q:
        conditions.append(AmbulatoryRecord.journal_number.ilike(f'%{escape_like(journal_number_q)}%', escape='\\'))
    if diagnosis_q:
        conditions.append(AmbulatoryRecord.diagnosis.ilike(f'%{escape_like(diagnosis_q)}%', escape='\\'))

    if conditions:
        q = q.filter(*conditions)

    # Dropdown data (cached)
    statuses = get_distinct_ambulatory_statuses()
    doctors = get_distinct_ambulatory_doctors()

    # Sorting
    sort_by = request.args.get('sort_by', 'date')
    sort_order = request.args.get('sort_order', 'desc')

    valid_columns = {
        'id': AmbulatoryRecord.id,
        'journal_number': AmbulatoryRecord.journal_number,
        'date': AmbulatoryRecord.date,
        'full_name': AmbulatoryRecord.full_name,
        'birth_date': AmbulatoryRecord.birth_date,
        'doctor': AmbulatoryRecord.doctor,
        'diagnosis': AmbulatoryRecord.diagnosis,
        'discharge_status': AmbulatoryRecord.discharge_status,
        'comment': AmbulatoryRecord.comment,
        'created_at': AmbulatoryRecord.created_at,
        'updated_at': AmbulatoryRecord.updated_at,
    }

    if sort_by in valid_columns:
        col = valid_columns[sort_by]
        if sort_order == 'asc':
            if sort_by in ('journal_number', 'full_name', 'doctor', 'diagnosis', 'discharge_status'):
                q = q.order_by(func.lower(col).asc())
            else:
                q = q.order_by(col.asc())
        else:
            if sort_by in ('journal_number', 'full_name', 'doctor', 'diagnosis', 'discharge_status'):
                q = q.order_by(func.lower(col).desc())
            else:
                q = q.order_by(col.desc())

    # Stats calculations
    count = q.count()
    count_discharged = q.filter(AmbulatoryRecord.discharge_status == STATUS_DISCHARGED).count()
    count_processing = q.filter(AmbulatoryRecord.discharge_status == STATUS_PROCESSING).count()
    count_violations = q.filter(AmbulatoryRecord.discharge_status == STATUS_VIOLATIONS).count()
    count_urgent = q.filter(AmbulatoryRecord.is_urgent == True).count()

    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 100, type=int)
    per_page = max(10, min(per_page, 200))

    pagination = q.paginate(page=page, per_page=per_page, error_out=False)
    records = pagination.items

    user_map = get_user_map()
    month_filter_value = f"{selected_year:04d}-{selected_month:02d}" if selected_year and selected_month else ""

    active_filters_count = (
        (1 if selected_status else 0) +
        (1 if selected_doctor else 0) +
        (1 if full_name_q else 0) +
        (1 if journal_number_q else 0) +
        (1 if diagnosis_q else 0)
    )

    return render_template('ambulatory_list.html',
                           records=records,
                           pagination=pagination,
                           statuses=statuses,
                           doctors=doctors,
                           selected_status=selected_status,
                           selected_doctor=selected_doctor,
                           full_name_q=full_name_q,
                           journal_number_q=journal_number_q,
                           diagnosis_q=diagnosis_q,
                           show_all=show_all,
                           selected_month=selected_month,
                           selected_year=selected_year,
                           month_filter_value=month_filter_value,
                           sort_by=sort_by,
                           sort_order=sort_order,
                           user_map=user_map,
                           count=count,
                           count_discharged=count_discharged,
                           count_processing=count_processing,
                           count_violations=count_violations,
                           count_urgent=count_urgent,
                           active_filters_count=active_filters_count)


@ambulatory_bp.route('/export', methods=['POST'])
@role_required('editor', 'viewer')
def export():
    """Export ambulatory records to Excel based on date filters."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    export_mode = request.form.get('export_mode', 'month').strip()

    if export_mode == 'range':
        from_str = request.form.get('from_date', '').strip()
        to_str = request.form.get('to_date', '').strip()
        if not from_str or not to_str:
            flash('Будь ласка, вкажіть обидві дати для експорту', 'warning')
            return redirect(url_for('ambulatory.index'))
        try:
            from_d = datetime.strptime(from_str, '%Y-%m-%d').date()
            to_d = datetime.strptime(to_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Невірний формат дати', 'warning')
            return redirect(url_for('ambulatory.index'))
        if from_d > to_d:
            flash('Дата "з" не може бути пізніше дати "по"', 'warning')
            return redirect(url_for('ambulatory.index'))
        conditions = [
            AmbulatoryRecord.date >= from_d,
            AmbulatoryRecord.date <= to_d,
        ]
    else:
        month_input = request.form.get('month_filter', '').strip()
        if not month_input:
            flash('Будь ласка, вкажіть місяць для експорту', 'warning')
            return redirect(url_for('ambulatory.index'))
        try:
            parts = month_input.split('-')
            if len(parts) != 2:
                raise ValueError()
            year = int(parts[0])
            month = int(parts[1])
            if not (1 <= month <= 12):
                raise ValueError()
            from_d = datetime(year, month, 1).date()
            if month == 12:
                to_d = datetime(year + 1, 1, 1).date()
            else:
                to_d = datetime(year, month + 1, 1).date()
            to_d = to_d - timedelta(days=1)
            conditions = [
                AmbulatoryRecord.date >= from_d,
                AmbulatoryRecord.date <= to_d,
            ]
        except ValueError:
            flash('Невірний формат місяця (очікується YYYY-MM)', 'warning')
            return redirect(url_for('ambulatory.index'))

    # Optional filters
    discharge_status = request.form.get('discharge_status', '').strip()
    doctor = request.form.get('doctor', '').strip()
    full_name_q = request.form.get('full_name', '').strip()
    journal_number_q = request.form.get('journal_number', '').strip()
    diagnosis_q = request.form.get('diagnosis', '').strip()

    if discharge_status:
        conditions.append(AmbulatoryRecord.discharge_status == discharge_status)
    if doctor:
        conditions.append(AmbulatoryRecord.doctor == doctor)
    if full_name_q:
        conditions.append(AmbulatoryRecord.full_name.ilike(f'%{escape_like(full_name_q)}%', escape='\\'))
    if journal_number_q:
        conditions.append(AmbulatoryRecord.journal_number.ilike(f'%{escape_like(journal_number_q)}%', escape='\\'))
    if diagnosis_q:
        conditions.append(AmbulatoryRecord.diagnosis.ilike(f'%{escape_like(diagnosis_q)}%', escape='\\'))

    q = AmbulatoryRecord.query.filter(*conditions)
    records = q.order_by(AmbulatoryRecord.date.desc()).all()

    if not records:
        flash('Записів не знайдено для експорту', 'warning')
        return redirect(url_for('ambulatory.index'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Амбулаторна допомога'

    headers = [
        'ID', 'Номер у журналі', 'Дата', 'П.І.П (повністю)', 'Дата народження',
        'Лікар', 'Діагноз', 'Статус виписки', 'Коментар',
        'Створено', 'Оновлено', 'Автор', 'Редактор'
    ]

    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color='1f4e78', end_color='1f4e78', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    user_map = get_user_map()

    # Data rows
    for r in records:
        row = [
            r.id,
            r.journal_number,
            r.date.strftime('%d.%m.%Y') if r.date else '',
            r.full_name,
            r.birth_date.strftime('%d.%m.%Y') if r.birth_date else '',
            r.doctor,
            r.diagnosis,
            r.discharge_status or '',
            r.comment or '',
            r.created_at.strftime('%d.%m.%Y %H:%M') if r.created_at else '',
            r.updated_at.strftime('%d.%m.%Y %H:%M') if r.updated_at else '',
            user_map.get(r.created_by, ''),
            user_map.get(r.updated_by, '')
        ]
        ws.append(row)

    # Auto-size columns
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        column = get_column_letter(i)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Save to BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    # Generate filename
    if export_mode == 'range':
        filename = f"ambulatory_export_{from_d.strftime('%d-%m-%Y')}_{to_d.strftime('%d-%m-%Y')}.xlsx"
        log_details = f'from={from_d} to={to_d} status={discharge_status} count={len(records)}'
    else:
        filename = f"ambulatory_export_{from_d.strftime('%m-%Y')}.xlsx"
        log_details = f'month={from_d.strftime("%m-%Y")} status={discharge_status} count={len(records)}'

    try:
        log_action(current_user.id, 'ambulatory.export', 'ambulatory_record', None, log_details)
        db.session.commit()
    except Exception:
        current_app.logger.exception('Failed to write audit log for ambulatory export')

    return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@ambulatory_bp.route('/print', methods=['POST'])
@role_required('editor', 'viewer')
def print_records():
    """Generate PDF print for ambulatory records with filters."""
    try:
        from_str = request.form.get('from_date', '').strip()
        to_str = request.form.get('to_date', '').strip()
        if not from_str or not to_str:
            flash('Будь ласка, вкажіть обидві дати для друку', 'warning')
            return redirect(url_for('ambulatory.index'))
        from_d = datetime.strptime(from_str, '%Y-%m-%d').date()
        to_d = datetime.strptime(to_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Невірний формат дати', 'warning')
        return redirect(url_for('ambulatory.index'))

    if from_d > to_d:
        flash('Дата "з" не може бути пізніше дати "по"', 'warning')
        return redirect(url_for('ambulatory.index'))

    # Get optional filters
    discharge_status = request.form.get('discharge_status', '').strip()
    doctor = request.form.get('doctor', '').strip()
    full_name_q = request.form.get('full_name', '').strip()
    journal_number_q = request.form.get('journal_number', '').strip()
    diagnosis_q = request.form.get('diagnosis', '').strip()

    # Build query
    conditions = [
        AmbulatoryRecord.date >= from_d,
        AmbulatoryRecord.date <= to_d,
    ]
    if discharge_status:
        conditions.append(AmbulatoryRecord.discharge_status == discharge_status)
    if doctor:
        conditions.append(AmbulatoryRecord.doctor == doctor)
    if full_name_q:
        conditions.append(AmbulatoryRecord.full_name.ilike(f'%{escape_like(full_name_q)}%', escape='\\'))
    if journal_number_q:
        conditions.append(AmbulatoryRecord.journal_number.ilike(f'%{escape_like(journal_number_q)}%', escape='\\'))
    if diagnosis_q:
        conditions.append(AmbulatoryRecord.diagnosis.ilike(f'%{escape_like(diagnosis_q)}%', escape='\\'))

    q = AmbulatoryRecord.query.filter(*conditions)
    records = q.order_by(AmbulatoryRecord.date.desc()).all()

    if not records:
        flash('Записів не знайдено для друку', 'warning')
        return redirect(url_for('ambulatory.index'))

    user_map = get_user_map()

    # Render template for print
    html_string = render_template('print_ambulatory.html',
                                 records=records,
                                 from_date=from_d,
                                 to_date=to_d,
                                 discharge_status=discharge_status,
                                 doctor=doctor,
                                 user_map=user_map,
                                 generated_by=current_user.username,
                                 generated_at=datetime.now(timezone(timedelta(hours=2))))

    try:
        from weasyprint import HTML
    except ImportError:
        flash('Для друку потрібен пакет WeasyPrint', 'danger')
        return redirect(url_for('ambulatory.index'))

    pdf = HTML(string=html_string).write_pdf()

    bio = BytesIO(pdf)
    bio.seek(0)

    try:
        log_details = f'from={from_d} to={to_d} status={discharge_status} count={len(records)}'
        log_action(current_user.id, 'ambulatory.print', 'ambulatory_record', None, log_details)
        db.session.commit()
    except Exception:
        current_app.logger.exception('Failed to write audit log for ambulatory print')

    filename = f"ambulatory_print_{datetime.now().strftime('%d-%m-%Y')}.pdf"
    return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/pdf')


@ambulatory_bp.route('/add', methods=['GET', 'POST'])
@role_required('operator', 'ambulatory')
def add_record():
    if request.method == 'POST':
        data, error = validate_ambulatory_form(request.form)
        if error:
            flash(error, 'warning')
            return redirect(url_for('ambulatory.add_record'))

        r = AmbulatoryRecord(
            journal_number=data['journal_number'],
            date=data['date'],
            full_name=data['full_name'],
            birth_date=data['birth_date'],
            doctor=data['doctor'],
            diagnosis=data['diagnosis'],
            discharge_status=STATUS_PROCESSING,
            comment=data['comment'],
            is_urgent=data['is_urgent'],
            created_by=current_user.id,
            updated_by=current_user.id
        )
        db.session.add(r)
        db.session.flush()

        log_action(current_user.id, 'ambulatory_record.create', 'ambulatory_record', r.id, f'full_name={r.full_name}')
        db.session.commit()

        clear_dropdown_cache()
        current_app.logger.info(f'AmbulatoryRecord created: {r.id} by {current_user.username}')
        flash(f'Запис "{r.full_name}" успішно додано', 'success')

        params = {}
        for k in ('discharge_status', 'doctor', 'full_name', 'journal_number', 'diagnosis'):
            v = request.form.get(f'filter_{k}', '').strip()
            if v:
                params[k] = v
        return redirect(url_for('ambulatory.index', **params))

    doctors = get_distinct_ambulatory_doctors()
    return render_template('ambulatory_add.html',
                           selected_status=request.args.get('discharge_status', ''),
                           selected_doctor=request.args.get('doctor', ''),
                           full_name_q=request.args.get('full_name', ''),
                           doctors=doctors)


@ambulatory_bp.route('/api/add', methods=['POST'])
@role_required('operator', 'ambulatory')
def api_add_record():
    """AJAX endpoint for adding records with support for 'save and add another'"""
    current_app.logger.info(f'API ambulatory/add called by {current_user.username}')

    data, error = validate_ambulatory_form(request.form)
    if error:
        return jsonify({'success': False, 'error': error}), 400

    r = AmbulatoryRecord(
        journal_number=data['journal_number'],
        date=data['date'],
        full_name=data['full_name'],
        birth_date=data['birth_date'],
        doctor=data['doctor'],
        diagnosis=data['diagnosis'],
        discharge_status=STATUS_PROCESSING,
        comment=data['comment'],
        is_urgent=data['is_urgent'],
        created_by=current_user.id,
        updated_by=current_user.id
    )

    try:
        db.session.add(r)
        db.session.flush()
        log_action(current_user.id, 'ambulatory_record.create', 'ambulatory_record', r.id, f'full_name={r.full_name}')
        db.session.commit()

        clear_dropdown_cache()
        current_app.logger.info(f'AmbulatoryRecord created via AJAX: {r.id} by {current_user.username}')

        return jsonify({
            'success': True,
            'record_id': r.id,
            'full_name': r.full_name,
            'message': f'Запис "{r.full_name}" успішно додано'
        })
    except Exception:
        db.session.rollback()
        current_app.logger.exception('Failed to create ambulatory record via AJAX')
        return jsonify({'success': False, 'error': 'Помилка при збереженні запису'}), 500


@ambulatory_bp.route('/<int:record_id>/edit', methods=['GET', 'POST'])
@role_required('editor')
def edit_record(record_id):
    r = db.get_or_404(AmbulatoryRecord, record_id)

    if request.method == 'POST':
        data, error = validate_ambulatory_form(request.form, require_status=True)
        if error:
            flash(error, 'warning')
            return redirect(url_for('ambulatory.edit_record', record_id=record_id))

        r.journal_number = data['journal_number']
        r.date = data['date']
        r.full_name = data['full_name']
        r.birth_date = data['birth_date']
        r.doctor = data['doctor']
        r.diagnosis = data['diagnosis']
        r.discharge_status = data['discharge_status']
        r.comment = data['comment']
        r.is_urgent = data['is_urgent']
        r.updated_by = current_user.id
        r.updated_at = datetime.now(timezone.utc)

        try:
            log_action(current_user.id, 'ambulatory_record.update', 'ambulatory_record', r.id, f'full_name={r.full_name}')
            db.session.commit()
        except Exception:
            db.session.rollback()
            current_app.logger.exception('Failed to update ambulatory record')
            flash('Помилка при збереженні змін', 'danger')
            return redirect(url_for('ambulatory.edit_record', record_id=record_id))

        clear_dropdown_cache()
        current_app.logger.info(f'AmbulatoryRecord updated: {r.id} by {current_user.username}')
        flash(f'Запис #{r.id} ({r.full_name}) успішно оновлено', 'success')

        params = {}
        for k in ('discharge_status', 'doctor', 'full_name', 'journal_number', 'diagnosis'):
            v = request.form.get(f'filter_{k}', '').strip()
            if v:
                params[k] = v
        return redirect(url_for('ambulatory.index', **params, _anchor=f'record-{r.id}'))

    doctors = get_distinct_ambulatory_doctors()
    return render_template('ambulatory_edit.html',
                           r=r,
                           selected_status=request.args.get('discharge_status', ''),
                           selected_doctor=request.args.get('doctor', ''),
                           full_name_q=request.args.get('full_name', ''),
                           doctors=doctors)


@ambulatory_bp.route('/api/<int:record_id>/edit', methods=['POST'])
@role_required('editor')
def api_edit_record(record_id):
    """AJAX endpoint for editing ambulatory records"""
    current_app.logger.info(f'API ambulatory/edit called by {current_user.username} for record {record_id}')

    r = db.get_or_404(AmbulatoryRecord, record_id)

    data, error = validate_ambulatory_form(request.form, require_status=True)
    if error:
        return jsonify({'success': False, 'error': error}), 400

    r.journal_number = data['journal_number']
    r.date = data['date']
    r.full_name = data['full_name']
    r.birth_date = data['birth_date']
    r.doctor = data['doctor']
    r.diagnosis = data['diagnosis']
    r.discharge_status = data['discharge_status']
    r.comment = data['comment']
    r.is_urgent = data['is_urgent']
    r.updated_by = current_user.id
    r.updated_at = datetime.now(timezone.utc)

    try:
        log_action(current_user.id, 'ambulatory_record.update', 'ambulatory_record', r.id, f'full_name={r.full_name}')
        db.session.commit()

        clear_dropdown_cache()
        current_app.logger.info(f'AmbulatoryRecord updated via AJAX: {r.id} by {current_user.username}')

        return jsonify({
            'success': True,
            'record_id': r.id,
            'full_name': r.full_name,
            'message': f'Запис "{r.full_name}" успішно оновлено'
        })
    except Exception:
        db.session.rollback()
        current_app.logger.exception('Failed to update ambulatory record via AJAX')
        return jsonify({'success': False, 'error': 'Помилка при оновленні запису'}), 500


@ambulatory_bp.route('/<int:record_id>/delete', methods=['POST'])
@role_required('admin')
def delete_record(record_id):
    r = db.get_or_404(AmbulatoryRecord, record_id)
    saved_id = r.id
    saved_name = r.full_name

    db.session.delete(r)
    log_action(current_user.id, 'ambulatory_record.delete', 'ambulatory_record', saved_id, f'full_name={saved_name}')
    db.session.commit()

    clear_dropdown_cache()
    current_app.logger.info(f'AmbulatoryRecord deleted: {saved_id} by {current_user.username}')
    flash(f'Запис #{saved_id} ({saved_name}) видалено', 'danger')

    params = {}
    for k in ('discharge_status', 'doctor', 'full_name', 'journal_number', 'diagnosis'):
        v = request.form.get(k, '').strip()
        if v:
            params[k] = v
    return redirect(url_for('ambulatory.index', **params))
