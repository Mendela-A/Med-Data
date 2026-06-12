# app/blueprints/records/routes.py
"""
Records (Dashboard) routes - main application routes for medical records
"""

from flask import render_template, redirect, url_for, flash, request, current_app, send_file, jsonify
from flask_login import login_required, current_user
from datetime import datetime, timezone, timedelta
from io import BytesIO
from sqlalchemy.orm import joinedload
from sqlalchemy import func, case

from app.extensions import db
from models import Record, User, Department, log_action
from decorators import role_required
from utils import (parse_date, parse_integer, parse_numeric, clear_dropdown_cache,
                   get_user_map, escape_like, validate_record_form,
                   get_distinct_statuses, get_distinct_physicians, get_distinct_departments,
                   get_status_options, get_default_status)
from constants import STATUS_DISCHARGED, STATUS_PROCESSING, STATUS_VIOLATIONS
from . import records_bp


# Routes
@records_bp.route('/')
@login_required
def index():
    # Redirect ambulatory role to ambulatory index
    if current_user.role == 'ambulatory':
        return redirect(url_for('ambulatory.index'))

    # Redirect viewer to statistics page ONLY if no filters applied
    if current_user.role == 'viewer' and not request.args:
        return redirect(url_for('admin.admin_statistics'))

    # Use Kyiv timezone (UTC+2, or UTC+3 during DST) for correct month detection
    # Simple approach: use UTC+2 as base (covers most of the year)
    kyiv_tz = timezone(timedelta(hours=2))
    now = datetime.now(kyiv_tz)

    # support a toggle to show all months
    show_all = request.args.get('all_months', '').lower() in ('1', 'true', 'yes')

    # Allow explicit month/year selection via query params or HTML5 month input (YYYY-MM format)
    # Also support from_date / to_date params (YYYY-MM-DD) from statistics page
    month_input = request.args.get('month_filter', '').strip()
    from_date_input = request.args.get('from_date', '').strip()
    to_date_input = request.args.get('to_date', '').strip()
    selected_month = None
    selected_year = None

    try:
        if from_date_input and to_date_input:
            # Date range mode from statistics page
            from datetime import date as date_type
            fd = date_type.fromisoformat(from_date_input)
            td = date_type.fromisoformat(to_date_input)
            if fd > td:
                fd, td = td, fd
            start = datetime(fd.year, fd.month, fd.day)
            end = datetime(td.year, td.month, td.day) + timedelta(days=1)
            selected_year = fd.year
            selected_month = fd.month
        elif month_input:
            # Parse HTML5 month input format: YYYY-MM
            parts = month_input.split('-')
            if len(parts) == 2:
                selected_year = int(parts[0])
                selected_month = int(parts[1])
                if 1 <= selected_month <= 12:
                    start = datetime(selected_year, selected_month, 1)
                    if selected_month == 12:
                        end = datetime(selected_year + 1, 1, 1)
                    else:
                        end = datetime(selected_year, selected_month + 1, 1)
                else:
                    raise ValueError()
            else:
                raise ValueError()
        else:
            # Default to current month
            start = datetime(now.year, now.month, 1)
            selected_year = now.year
            selected_month = now.month
            if now.month == 12:
                end = datetime(now.year + 1, 1, 1)
            else:
                end = datetime(now.year, now.month + 1, 1)
    except Exception:
        # fallback to current month
        start = datetime(now.year, now.month, 1)
        selected_year = now.year
        selected_month = now.month
        if now.month == 12:
            end = datetime(now.year + 1, 1, 1)
        else:
            end = datetime(now.year, now.month + 1, 1)

    # base query (by default for current month, unless show_all)
    q = Record.query.options(joinedload(Record.creator), joinedload(Record.updater))
    date_conditions = []
    if not show_all:
        # show records discharged in the current month by date_of_discharge
        date_conditions = [
            Record.date_of_discharge != None,
            Record.date_of_discharge >= start.date(),
            Record.date_of_discharge < end.date(),
        ]
        q = q.filter(*date_conditions)

    # --- filtering from query params ---
    selected_status = request.args.get('discharge_status', '').strip()
    selected_physician = request.args.get('treating_physician', '').strip()
    selected_department = request.args.get('discharge_department', '').strip()
    history_q = request.args.get('history', '').strip()
    full_name_q = request.args.get('full_name', '').strip()
    has_death_date = request.args.get('has_death_date', '').lower() in ('1', 'true', 'yes')
    filter_history_submitted = request.args.get('history_submitted', '').strip()  # '1'=здані, '0'=незданні, ''=всі

    conditions = []
    if selected_status:
        conditions.append(Record.discharge_status == selected_status)
    if selected_physician:
        conditions.append(Record.treating_physician == selected_physician)
    if selected_department:
        conditions.append(Record.discharge_department == selected_department)
    if history_q:
        conditions.append(Record.history.like(f'%{escape_like(history_q)}%', escape='\\'))
    if full_name_q:
        conditions.append(Record.full_name.ilike(f'%{escape_like(full_name_q)}%', escape='\\'))
    if has_death_date:
        conditions.append(Record.date_of_death != None)
    if filter_history_submitted == '1':
        conditions.append(Record.history_submitted == True)
    elif filter_history_submitted == '0':
        conditions.append(Record.history_submitted == False)
    if conditions:
        q = q.filter(*conditions)

    # values for dropdowns (cached)
    statuses = get_distinct_statuses()
    physicians = get_distinct_physicians()
    departments = get_distinct_departments()

    # Sorting
    sort_by = request.args.get('sort_by', 'date_of_discharge')
    sort_order = request.args.get('sort_order', 'desc')

    valid_columns = {
        'id': Record.id,
        'date_of_discharge': Record.date_of_discharge,
        'full_name': Record.full_name,
        'discharge_department': Record.discharge_department,
        'treating_physician': Record.treating_physician,
        'history': Record.history,
        'k_days': Record.k_days,
        'discharge_status': Record.discharge_status,
        'date_of_death': Record.date_of_death,
        'created_at': Record.created_at,
        'updated_at': Record.updated_at,
    }

    if sort_by in valid_columns:
        col = valid_columns[sort_by]
        if sort_order == 'asc':
            # For string columns, use case-insensitive sorting
            if sort_by in ('full_name', 'discharge_department', 'treating_physician', 'history', 'discharge_status'):
                q = q.order_by(func.lower(col).asc())
            else:
                q = q.order_by(col.asc())
        else:
            if sort_by in ('full_name', 'discharge_department', 'treating_physician', 'history', 'discharge_status'):
                q = q.order_by(func.lower(col).desc())
            else:
                q = q.order_by(col.desc())

    # Calculate statistics counts BEFORE pagination
    # Total count
    count = q.count()

    # Count deceased (priority: any record with date_of_death)
    count_deceased = q.filter(Record.date_of_death != None).count()

    # Other counts: EXCLUDE records with date_of_death
    q_alive = q.filter(Record.date_of_death == None)
    count_discharged = q_alive.filter(Record.discharge_status == STATUS_DISCHARGED).count()
    count_processing = q_alive.filter(Record.discharge_status == STATUS_PROCESSING).count()
    count_violations = q_alive.filter(Record.discharge_status == STATUS_VIOLATIONS).count()

    # Лічильники ургентних і зданих (по всьому поточному filtered set)
    count_urgent = q.filter(Record.is_urgent == True).count()
    count_planned = q.filter(Record.is_urgent == False).count()
    count_submitted = q.filter(Record.history_submitted == True).count()
    count_not_submitted = q.filter(Record.history_submitted == False).count()

    # Довідник статусів: селекти/бейджі + динамічні піли для несистемних
    # статусів (рахуються за тим самим правилом — без дати смерті)
    status_defs = get_status_options('records')
    status_meta = {s['name']: s for s in get_status_options('records', include_inactive=True)}
    extra_status_counts = {}
    if any(s['show_in_stats'] and not s['is_system'] for s in status_defs):
        counts_q = db.session.query(Record.discharge_status, func.count(Record.id)) \
            .filter(Record.date_of_death == None)
        if date_conditions:
            counts_q = counts_q.filter(*date_conditions)
        if conditions:
            counts_q = counts_q.filter(*conditions)
        extra_status_counts = {name: cnt for name, cnt in
                               counts_q.group_by(Record.discharge_status).all() if name}

    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 100, type=int)
    per_page = max(10, min(per_page, 200))

    pagination = q.paginate(page=page, per_page=per_page, error_out=False)
    records = pagination.items

    # user mapping for created_by / updated_by
    user_map = get_user_map()

    # Format month_filter_value for HTML5 month input (YYYY-MM)
    month_filter_value = f"{selected_year:04d}-{selected_month:02d}" if selected_year and selected_month else ""

    return render_template('dashboard.html',
                          records=records,
                          pagination=pagination,
                          statuses=statuses,
                          status_defs=status_defs,
                          status_meta=status_meta,
                          extra_status_counts=extra_status_counts,
                          default_status=get_default_status('records'),
                          physicians=physicians,
                          departments=departments,
                          selected_status=selected_status,
                          selected_physician=selected_physician,
                          selected_department=selected_department,
                          history_q=history_q,
                          full_name_q=full_name_q,
                          has_death_date=has_death_date,
                          show_all=show_all,
                          selected_month=selected_month,
                          selected_year=selected_year,
                          month_filter_value=month_filter_value,
                          sort_by=sort_by,
                          sort_order=sort_order,
                          user_map=user_map,
                          count=count,
                          count_discharged=count_discharged,
                          count_processing=count_processing,
                          count_violations=count_violations,
                          count_deceased=count_deceased,
                          filter_history_submitted=filter_history_submitted,
                          count_urgent=count_urgent,
                          count_planned=count_planned,
                          count_submitted=count_submitted,
                          count_not_submitted=count_not_submitted,
                          active_filters_count=(1 if selected_status else 0) + (1 if selected_physician else 0) + (1 if selected_department else 0) + (1 if history_q else 0) + (1 if full_name_q else 0) + (1 if filter_history_submitted else 0))


@records_bp.route('/export', methods=['POST'])
@role_required('editor', 'viewer')
def export():
    """Export records to Excel based on form data (month or date range)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    export_mode = request.form.get('export_mode', 'month').strip()

    if export_mode == 'range':
        # date range mode
        from_str = request.form.get('from_date', '').strip()
        to_str = request.form.get('to_date', '').strip()
        if not from_str or not to_str:
            flash('Будь ласка, вкажіть обидві дати для експорту', 'warning')
            return redirect(url_for('records.index'))
        try:
            from_d = datetime.strptime(from_str, '%Y-%m-%d').date()
            to_d = datetime.strptime(to_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Невірний формат дати', 'warning')
            return redirect(url_for('records.index'))
        if from_d > to_d:
            flash('Дата "з" не може бути пізніше дати "по"', 'warning')
            return redirect(url_for('records.index'))
        # query
        conditions = [
            Record.date_of_discharge != None,
            Record.date_of_discharge >= from_d,
            Record.date_of_discharge <= to_d,
        ]
    else:
        # month mode (default)
        month_input = request.form.get('month_filter', '').strip()
        if not month_input:
            flash('Будь ласка, вкажіть місяць для експорту', 'warning')
            return redirect(url_for('records.index'))
        try:
            parts = month_input.split('-')
            if len(parts) != 2:
                raise ValueError()
            year = int(parts[0])
            month = int(parts[1])
            if not (1 <= month <= 12):
                raise ValueError()
            from_d = datetime(year, month, 1).date()
            if month == 12:
                to_d = datetime(year + 1, 1, 1).date()
            else:
                to_d = datetime(year, month + 1, 1).date()
            # adjust to_d so it's inclusive (last day of month)
            to_d = to_d - timedelta(days=1)
            conditions = [
                Record.date_of_discharge != None,
                Record.date_of_discharge >= from_d,
                Record.date_of_discharge <= to_d,
            ]
        except ValueError:
            flash('Невірний формат місяця (очікується YYYY-MM)', 'warning')
            return redirect(url_for('records.index'))

    # Optional filters
    discharge_status = request.form.get('discharge_status', '').strip()
    treating_physician = request.form.get('treating_physician', '').strip()
    discharge_department = request.form.get('discharge_department', '').strip()
    history_q = request.form.get('history', '').strip()
    full_name_q = request.form.get('full_name', '').strip()

    if discharge_status:
        conditions.append(Record.discharge_status == discharge_status)
    if treating_physician:
        conditions.append(Record.treating_physician == treating_physician)
    if discharge_department:
        conditions.append(Record.discharge_department == discharge_department)
    if history_q:
        conditions.append(Record.history.like(f'%{escape_like(history_q)}%', escape='\\'))
    if full_name_q:
        conditions.append(Record.full_name.ilike(f'%{escape_like(full_name_q)}%', escape='\\'))

    q = Record.query.filter(*conditions)
    records = q.order_by(Record.date_of_discharge.desc()).all()

    if not records:
        flash('Записів не знайдено для експорту', 'warning')
        return redirect(url_for('records.index'))

    # Determine access level
    use_write_only = current_user.role == 'viewer'

    # user mapping for creator/updater names
    user_map = get_user_map()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Записи'

    # Headers
    if use_write_only:
        headers = ['ID', 'Дата виписки', 'ПІБ', 'Відділення', 'Лікар', 'Історія хвороби', 'К днів', 'Статус виписки']
    else:
        headers = ['ID', 'Дата виписки', 'ПІБ', 'Відділення', 'Лікар', 'Історія хвороби', 'К днів', 'Статус виписки', 'АДСЖ', 'Сума', 'Дата смерті', 'Коментар', 'Створено', 'Оновлено', 'Автор', 'Редактор']

    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows
    for r in records:
        if use_write_only:
            row = [
                r.id,
                r.date_of_discharge.strftime('%d.%m.%Y') if r.date_of_discharge else '',
                r.full_name,
                r.discharge_department or '',
                r.treating_physician,
                r.history,
                r.k_days,
                r.discharge_status or ''
            ]
        else:
            row = [
                r.id,
                r.date_of_discharge.strftime('%d.%m.%Y') if r.date_of_discharge else '',
                r.full_name,
                r.discharge_department or '',
                r.treating_physician,
                r.history,
                r.k_days,
                r.discharge_status or '',
                r.adsj or '',
                f"{int(r.suma):,}".replace(",", " ") if r.suma is not None else '',
                r.date_of_death.strftime('%d.%m.%Y') if r.date_of_death else '',
                r.comment or '',
                r.created_at.strftime('%d.%m.%Y %H:%M') if r.created_at else '',
                r.updated_at.strftime('%d.%m.%Y %H:%M') if r.updated_at else '',
                user_map.get(r.created_by, ''),
                user_map.get(r.updated_by, '')
            ]
        ws.append(row)

    # Auto-size columns
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        column = get_column_letter(i)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Save to BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    # Generate filename
    if export_mode == 'range':
        filename = f"vipiski_export_{from_d.strftime('%d-%m-%Y')}_{to_d.strftime('%d-%m-%Y')}.xlsx"
        log_details = f'from={from_d} to={to_d} status={discharge_status} count={len(records)}'
    else:
        filename = f"vipiski_export_{from_d.strftime('%m-%Y')}.xlsx"
        log_details = f'month={from_d.strftime("%m-%Y")} status={discharge_status} count={len(records)}'

    # Audit log
    try:
        log_action(current_user.id, 'records.export', 'export', None, log_details)
        db.session.commit()
    except Exception:
        current_app.logger.exception('Failed to write audit log for export')
    current_app.logger.info(f'Export by {getattr(current_user, "username", "unknown")}: {log_details} write_only={use_write_only}')

    return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@records_bp.route('/records/print', methods=['POST'])
@role_required('editor', 'viewer')
def print_records():
    """Generate PDF print for records with filters"""
    try:
        from_str = request.form.get('from_date', '').strip()
        to_str = request.form.get('to_date', '').strip()
        if not from_str or not to_str:
            flash('Будь ласка, вкажіть обидві дати для друку', 'warning')
            return redirect(url_for('records.index'))
        from_d = datetime.strptime(from_str, '%Y-%m-%d').date()
        to_d = datetime.strptime(to_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Невірний формат дати', 'warning')
        return redirect(url_for('records.index'))

    if from_d > to_d:
        flash('Дата "з" не може бути пізніше дати "по"', 'warning')
        return redirect(url_for('records.index'))

    # Get optional filters
    discharge_status = request.form.get('discharge_status', '').strip()
    treating_physician = request.form.get('treating_physician', '').strip()
    discharge_department = request.form.get('discharge_department', '').strip()
    history_q = request.form.get('history', '').strip()
    full_name_q = request.form.get('full_name', '').strip()

    # Build query with filters
    conditions = [
        Record.date_of_discharge != None,
        Record.date_of_discharge >= from_d,
        Record.date_of_discharge <= to_d,
    ]
    if discharge_status:
        conditions.append(Record.discharge_status == discharge_status)
    if treating_physician:
        conditions.append(Record.treating_physician == treating_physician)
    if discharge_department:
        conditions.append(Record.discharge_department == discharge_department)
    if history_q:
        conditions.append(Record.history.like(f'%{escape_like(history_q)}%', escape='\\'))
    if full_name_q:
        conditions.append(Record.full_name.ilike(f'%{escape_like(full_name_q)}%', escape='\\'))

    q = Record.query.filter(*conditions)
    records = q.order_by(Record.date_of_discharge.desc()).all()

    if not records:
        flash('Записів не знайдено для друку', 'warning')
        return redirect(url_for('records.index'))

    # Get user mapping
    user_map = get_user_map()

    # Render HTML template for print
    html_string = render_template('print_records.html',
                                 records=records,
                                 from_date=from_d,
                                 to_date=to_d,
                                 discharge_status=discharge_status,
                                 treating_physician=treating_physician,
                                 discharge_department=discharge_department,
                                 user_map=user_map,
                                 generated_by=current_user.username,
                                 generated_at=datetime.now(timezone(timedelta(hours=2))))

    # Generate PDF
    try:
        from weasyprint import HTML, CSS
    except ImportError:
        flash('Для друку потрібен пакет WeasyPrint', 'danger')
        return redirect(url_for('records.index'))

    pdf = HTML(string=html_string).write_pdf()

    # Create BytesIO object
    bio = BytesIO(pdf)
    bio.seek(0)

    # Log action
    try:
        log_details = f'from={from_d} to={to_d} status={discharge_status} count={len(records)}'
        log_action(current_user.id, 'records.print', 'print', None, log_details)
        db.session.commit()
    except Exception:
        current_app.logger.exception('Failed to write audit log for records.print')

    filename = f"vipiski_print_{datetime.now().strftime('%d-%m-%Y')}.pdf"
    return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/pdf')


@records_bp.route('/records/add', methods=['GET', 'POST'])
@role_required('operator')
def add_record():

    if request.method == 'POST':
        data, error = validate_record_form(request.form)
        if error:
            flash(error, 'warning')
            return redirect(url_for('records.add_record'))

        r = Record(
            date_of_discharge=data['date_of_discharge'],
            full_name=data['full_name'],
            discharge_department=data['discharge_department'],
            treating_physician=data['treating_physician'],
            history=data['history'],
            k_days=data['k_days'],
            discharge_status=get_default_status('records'),
            date_of_death=data['date_of_death'],
            comment=data['comment'],
            is_urgent=data['is_urgent'] if current_user.role in ('operator', 'admin') else None,
            history_submitted=data['history_submitted'] if current_user.role in ('operator', 'admin') else False,
            created_by=current_user.id,
            updated_by=current_user.id
        )
        db.session.add(r)
        db.session.flush()  # assigns r.id
        log_action(current_user.id, 'record.create', 'record', r.id, f'full_name={r.full_name}')
        db.session.commit()
        # Clear dropdown cache so newly added values appear in dropdowns
        clear_dropdown_cache()
        current_app.logger.info(f'Record created: {r.id} by {current_user.username}')
        flash(f'Запис "{r.full_name}" успішно додано', 'success')
        # preserve filters from form (if any)
        params = {}
        for k in ('discharge_status', 'treating_physician', 'history'):
            v = request.form.get(f'filter_{k}', '').strip()
            if v:
                params[k] = v
        if request.form.get('filter_has_death_date', '').strip():
            params['has_death_date'] = '1'
        return redirect(url_for('records.index', **params))

    # GET: pass through any filters so add form can include hidden fields and departments
    departments = Department.query.order_by(Department.name).all()
    # Get distinct physicians for autocomplete (cached)
    physicians = get_distinct_physicians()
    return render_template('add_record.html', selected_status=request.args.get('discharge_status', ''), selected_physician=request.args.get('treating_physician', ''), history_q=request.args.get('history', ''), departments=departments, selected_department=request.args.get('discharge_department', ''), physicians=physicians)


@records_bp.route('/api/records/add', methods=['POST'])
@role_required('operator')
def api_add_record():
    """AJAX endpoint for adding records with support for 'save and add another'"""
    current_app.logger.info(f'API add_record called by {current_user.username}')

    data, error = validate_record_form(request.form)
    if error:
        return jsonify({'success': False, 'error': error}), 400

    r = Record(
        date_of_discharge=data['date_of_discharge'],
        full_name=data['full_name'],
        discharge_department=data['discharge_department'],
        treating_physician=data['treating_physician'],
        history=data['history'],
        k_days=data['k_days'],
        discharge_status=get_default_status('records'),
        date_of_death=data['date_of_death'],
        comment=data['comment'],
        is_urgent=data['is_urgent'],
        history_submitted=data['history_submitted'],
        created_by=current_user.id,
        updated_by=current_user.id
    )

    try:
        db.session.add(r)
        db.session.flush()  # assigns r.id
        log_action(current_user.id, 'record.create', 'record', r.id, f'full_name={r.full_name}')
        db.session.commit()

        # Clear dropdown cache
        clear_dropdown_cache()

        current_app.logger.info(f'Record created via AJAX: {r.id} by {current_user.username}')

        return jsonify({
            'success': True,
            'record_id': r.id,
            'full_name': r.full_name,
            'message': f'Запис "{r.full_name}" успішно додано'
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Failed to create record via AJAX')
        return jsonify({'success': False, 'error': 'Помилка при збереженні запису'}), 500


@records_bp.route('/api/records/<int:record_id>/status', methods=['POST'])
@role_required('operator')
def api_update_record_status(record_id):
    """AJAX endpoint for operator: update is_urgent and history_submitted only."""
    r = db.get_or_404(Record, record_id)

    is_urgent_raw = request.form.get('is_urgent', '')
    if is_urgent_raw == 'urgent':
        r.is_urgent = True
    elif is_urgent_raw == 'planned':
        r.is_urgent = False
    else:
        r.is_urgent = None

    r.history_submitted = request.form.get('history_submitted') == '1'
    r.updated_by = current_user.id
    r.updated_at = datetime.now(timezone.utc)

    try:
        log_action(current_user.id, 'record.update_status', 'record', r.id,
                   f'is_urgent={r.is_urgent} history_submitted={r.history_submitted}')
        db.session.commit()
        return jsonify({'success': True, 'record_id': r.id})
    except Exception:
        db.session.rollback()
        current_app.logger.exception('Failed to update record status')
        return jsonify({'success': False, 'error': 'Помилка при оновленні статусу'}), 500


@records_bp.route('/api/records/<int:record_id>/edit', methods=['POST'])
@role_required('editor')
def api_edit_record(record_id):
    """AJAX endpoint for editing records"""
    current_app.logger.info(f'API edit_record called by {current_user.username} for record {record_id}')

    r = db.get_or_404(Record, record_id)

    data, error = validate_record_form(request.form, require_status_and_dept=True)
    if error:
        return jsonify({'success': False, 'error': error}), 400

    r.date_of_discharge = data['date_of_discharge']
    r.full_name = data['full_name']
    r.discharge_department = data['discharge_department']
    r.treating_physician = data['treating_physician']
    r.history = data['history']
    r.k_days = data['k_days']
    r.discharge_status = data['discharge_status']
    r.date_of_death = data['date_of_death']
    r.comment = data['comment']
    r.adsj = data['adsj']
    r.suma = data['suma']
    if current_user.role in ('operator', 'admin'):
        r.is_urgent = data['is_urgent']
        r.history_submitted = data['history_submitted']
    r.updated_by = current_user.id
    r.updated_at = datetime.now(timezone.utc)

    try:
        log_action(current_user.id, 'record.update', 'record', r.id, f'full_name={r.full_name}')
        db.session.commit()

        # Clear dropdown cache
        clear_dropdown_cache()

        current_app.logger.info(f'Record updated via AJAX: {r.id} by {current_user.username}')

        return jsonify({
            'success': True,
            'record_id': r.id,
            'full_name': r.full_name,
            'message': f'Запис "{r.full_name}" успішно оновлено'
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Failed to update record via AJAX')
        return jsonify({'success': False, 'error': 'Помилка при оновленні запису'}), 500


@records_bp.route('/records/<int:record_id>/edit', methods=['GET', 'POST'])
@role_required('editor')
def edit_record(record_id):

    r = db.get_or_404(Record, record_id)

    if request.method == 'POST':
        data, error = validate_record_form(request.form, require_status_and_dept=True)
        if error:
            flash(error, 'warning')
            return redirect(url_for('records.edit_record', record_id=record_id))

        r.date_of_discharge = data['date_of_discharge']
        r.full_name = data['full_name']
        r.discharge_department = data['discharge_department']
        r.treating_physician = data['treating_physician']
        r.history = data['history']
        r.k_days = data['k_days']
        r.discharge_status = data['discharge_status']
        r.date_of_death = data['date_of_death']
        r.comment = data['comment']
        r.adsj = data['adsj']
        r.suma = data['suma']
        r.is_urgent = data['is_urgent']
        r.history_submitted = data['history_submitted']
        r.updated_by = current_user.id
        r.updated_at = datetime.now(timezone.utc)

        try:
            log_action(current_user.id, 'record.update', 'record', r.id, f'full_name={r.full_name}')
            db.session.commit()
        except Exception:
            db.session.rollback()
            current_app.logger.exception('Failed to update record')
            flash('Помилка при збереженні змін', 'danger')
            return redirect(url_for('records.edit_record', record_id=record_id))
        # Clear dropdown cache after editing record
        clear_dropdown_cache()
        current_app.logger.info(f'Record updated: {r.id} by {current_user.username}')
        flash(f'Запис #{r.id} ({r.full_name}) успішно оновлено', 'success')
        params = {}
        for k in ('discharge_status', 'treating_physician', 'history'):
            v = request.form.get(f'filter_{k}', '').strip()
            if v:
                params[k] = v
        if request.form.get('filter_has_death_date', '').strip():
            params['has_death_date'] = '1'
        # Add anchor to scroll to edited record
        return redirect(url_for('records.index', **params, _anchor=f'record-{r.id}'))

    # GET -> render form with record data (pass filters through if present) and departments
    departments = Department.query.order_by(Department.name).all()
    # Get distinct physicians for autocomplete (cached)
    physicians = get_distinct_physicians()
    return render_template('edit_record.html', r=r, status_defs=get_status_options('records'), selected_status=request.args.get('discharge_status', ''), selected_physician=request.args.get('treating_physician', ''), history_q=request.args.get('history', ''), departments=departments, physicians=physicians)


@records_bp.route('/records/<int:record_id>/delete', methods=['POST'])
@role_required('admin')
def delete_record(record_id):
    r = db.get_or_404(Record, record_id)
    saved_id = r.id
    saved_name = r.full_name
    db.session.delete(r)
    log_action(current_user.id, 'record.delete', 'record', saved_id, f'full_name={saved_name}')
    db.session.commit()
    # Clear dropdown cache after deleting record
    clear_dropdown_cache()
    current_app.logger.info(f'Record deleted: {saved_id} by {current_user.username}')
    flash(f'Запис #{saved_id} ({saved_name}) видалено', 'danger')
    # preserve filters from form (if any)
    params = {}
    for k in ('discharge_status', 'treating_physician', 'discharge_department', 'history', 'full_name'):
        v = request.form.get(k, '').strip()
        if v:
            params[k] = v
    if request.form.get('has_death_date', '').strip():
        params['has_death_date'] = '1'
    return redirect(url_for('records.index', **params))


@records_bp.route('/records/report/submission', methods=['POST'])
@role_required('operator', 'editor', 'viewer')
def report_submission():
    from_str = request.form.get('from_date', '').strip()
    to_str = request.form.get('to_date', '').strip()
    if not from_str or not to_str:
        flash('Будь ласка, вкажіть обидві дати', 'warning')
        return redirect(url_for('records.index'))
    try:
        from_d = datetime.strptime(from_str, '%Y-%m-%d').date()
        to_d = datetime.strptime(to_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Невірний формат дати', 'warning')
        return redirect(url_for('records.index'))
    if from_d > to_d:
        from_d, to_d = to_d, from_d
    query_end = to_d + timedelta(days=1)

    filter_physician = request.form.get('treating_physician', '').strip()

    base_filter = [
        Record.date_of_discharge.isnot(None),
        Record.date_of_discharge >= from_d,
        Record.date_of_discharge < query_end,
        Record.treating_physician.isnot(None),
    ]
    if filter_physician:
        base_filter.append(Record.treating_physician == filter_physician)

    submission_row = db.session.query(
        func.sum(case((Record.history_submitted == True, 1), else_=0)).label('submitted'),
        func.sum(case((Record.history_submitted == False, 1), else_=0)).label('not_submitted'),
    ).filter(*[
        Record.date_of_discharge.isnot(None),
        Record.date_of_discharge >= from_d,
        Record.date_of_discharge < query_end,
    ] + ([Record.treating_physician == filter_physician] if filter_physician else [])).first()

    submission_by_physician = db.session.query(
        Record.treating_physician,
        func.sum(case((Record.history_submitted == True, 1), else_=0)).label('submitted'),
        func.sum(case((Record.history_submitted == False, 1), else_=0)).label('not_submitted'),
        func.count(Record.id).label('total')
    ).filter(*base_filter).group_by(Record.treating_physician).order_by(
        func.sum(case((Record.history_submitted == False, 1), else_=0)).desc()
    ).all()

    kyiv_tz = timezone(timedelta(hours=2))
    try:
        from weasyprint import HTML
    except ImportError:
        flash('Для формування PDF потрібен пакет WeasyPrint', 'danger')
        return redirect(url_for('records.index'))

    html_string = render_template(
        'print_submission.html',
        from_date=from_d,
        to_date=to_d,
        filter_physician=filter_physician,
        submission_submitted=submission_row.submitted or 0,
        submission_not_submitted=submission_row.not_submitted or 0,
        submission_by_physician=submission_by_physician,
        generated_by=current_user.username,
        generated_at=datetime.now(kyiv_tz),
    )
    pdf = HTML(string=html_string).write_pdf()
    bio = BytesIO(pdf)
    bio.seek(0)
    filename = f"submission_report_{from_d.strftime('%d-%m-%Y')}_{to_d.strftime('%d-%m-%Y')}.pdf"
    try:
        log_action(current_user.id, 'records.report_submission', 'report', None,
                   f'from={from_d} to={to_d} physician={filter_physician}')
        db.session.commit()
    except Exception:
        current_app.logger.exception('Failed to log report_submission action')
    return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/pdf')


@records_bp.route('/records/report/urgency', methods=['POST'])
@role_required('operator', 'editor', 'viewer')
def report_urgency():
    from_str = request.form.get('from_date', '').strip()
    to_str = request.form.get('to_date', '').strip()
    if not from_str or not to_str:
        flash('Будь ласка, вкажіть обидві дати', 'warning')
        return redirect(url_for('records.index'))
    try:
        from_d = datetime.strptime(from_str, '%Y-%m-%d').date()
        to_d = datetime.strptime(to_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Невірний формат дати', 'warning')
        return redirect(url_for('records.index'))
    if from_d > to_d:
        from_d, to_d = to_d, from_d
    query_end = to_d + timedelta(days=1)

    filter_department = request.form.get('discharge_department', '').strip()

    date_filter = [
        Record.date_of_discharge.isnot(None),
        Record.date_of_discharge >= from_d,
        Record.date_of_discharge < query_end,
    ]
    dept_filter = ([Record.discharge_department == filter_department] if filter_department else [])

    urgency_row = db.session.query(
        func.sum(case((Record.is_urgent == True, 1), else_=0)).label('urgent'),
        func.sum(case((Record.is_urgent == False, 1), else_=0)).label('planned'),
        func.sum(case((Record.is_urgent.is_(None), 1), else_=0)).label('unset'),
    ).filter(*(date_filter + dept_filter)).first()

    urgency_by_dept = db.session.query(
        Record.discharge_department,
        func.sum(case((Record.is_urgent == True, 1), else_=0)).label('urgent'),
        func.sum(case((Record.is_urgent == False, 1), else_=0)).label('planned'),
        func.sum(case((Record.is_urgent.is_(None), 1), else_=0)).label('unset'),
        func.count(Record.id).label('total')
    ).filter(*(date_filter + dept_filter + [Record.discharge_department.isnot(None)])
    ).group_by(Record.discharge_department).order_by(
        func.sum(case((Record.is_urgent == True, 1), else_=0)).desc()
    ).all()

    kyiv_tz = timezone(timedelta(hours=2))
    try:
        from weasyprint import HTML
    except ImportError:
        flash('Для формування PDF потрібен пакет WeasyPrint', 'danger')
        return redirect(url_for('records.index'))

    html_string = render_template(
        'print_urgency.html',
        from_date=from_d,
        to_date=to_d,
        filter_department=filter_department,
        urgency_urgent=urgency_row.urgent or 0,
        urgency_planned=urgency_row.planned or 0,
        urgency_unset=urgency_row.unset or 0,
        urgency_by_dept=urgency_by_dept,
        generated_by=current_user.username,
        generated_at=datetime.now(kyiv_tz),
    )
    pdf = HTML(string=html_string).write_pdf()
    bio = BytesIO(pdf)
    bio.seek(0)
    filename = f"urgency_report_{from_d.strftime('%d-%m-%Y')}_{to_d.strftime('%d-%m-%Y')}.pdf"
    try:
        log_action(current_user.id, 'records.report_urgency', 'report', None,
                   f'from={from_d} to={to_d} dept={filter_department}')
        db.session.commit()
    except Exception:
        current_app.logger.exception('Failed to log report_urgency action')
    return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/pdf')
